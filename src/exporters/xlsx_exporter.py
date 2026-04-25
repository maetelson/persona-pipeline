"""Final workbook exporter for service-fit persona outputs."""

from __future__ import annotations

import json
import re
import warnings
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analysis.workbook_bundle import validate_workbook_frames
from src.utils.io import ensure_dir, load_yaml
from src.utils.pipeline_schema import WORKBOOK_SHEET_NAMES, round_frame_ratios


ILLEGAL_EXCEL_CHAR_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, color="1F1F1F")
TITLE_FONT = Font(bold=True, size=12)
SUBTLE_FILL = PatternFill(fill_type="solid", fgColor="F3F6F9")
PERCENT_LITERAL_FORMAT = '0.0"%"'
RATIO_FORMAT = '0.00'
INTEGER_FORMAT = '0'
DECIMAL_FORMAT = '0.0'

DISPLAY_HEADER_OVERRIDES = {
    "share_of_core_labeled": "share_of_persona_core_labeled_pct",
    "share_of_all_labeled": "share_of_all_labeled_pct",
    "denominator_type": "denominator_type_key",
    "denominator_value": "denominator_row_count",
    "pct_of_persona": "pct_of_persona_rows",
    "grain": "row_grain",
    "metric_type": "metric_value_type",
    "selection_strength": "selected_example_strength",
    "grounding_strength": "example_grounding_strength",
}

DISPLAY_TO_RAW_HEADER_OVERRIDES = {value: key for key, value in DISPLAY_HEADER_OVERRIDES.items()}
DISPLAY_TO_RAW_HEADER_OVERRIDES.update(
    {
        "metric_key": "metric",
        "metric_value": "value",
        "threshold_rule": "threshold",
        "workbook_label": "",
        "display_label": "",
        "row_count": "count",
        "raw_record_rows_for_source": "raw_count",
        "normalized_post_rows_for_source": "normalized_count",
        "valid_candidate_rows_for_source": "valid_count",
        "prefiltered_valid_post_rows_for_source": "prefiltered_valid_count",
        "episode_rows_for_source": "episode_count",
        "labeled_episode_rows_for_source": "labeled_count",
        "share_of_labeled_episode_rows_pct": "share_of_labeled",
    }
)

OVERVIEW_METRIC_LABELS = {
    "persona_readiness_state": "Persona readiness state",
    "persona_readiness_label": "Workbook interpretation label",
    "persona_asset_class": "Workbook asset class",
    "persona_readiness_gate_status": "Persona readiness gate status",
    "persona_completion_claim_allowed": "Persona-complete claim allowed",
    "persona_usage_restriction": "Workbook usage restriction",
    "persona_readiness_summary": "Persona readiness summary",
    "persona_readiness_blockers": "Persona readiness blockers",
    "persona_readiness_rule": "Persona readiness policy rule",
    "promotion_constraint_status": "Promotion constraint status",
    "promotion_constraint_summary": "Promotion constraint summary",
    "source_action_priority_summary": "Highest-priority source actions",
    "overall_status": "Overall workbook status",
    "quality_flag": "Quality flag derived from overall status",
    "quality_flag_rule": "Quality flag decision rule",
    "composite_reason_keys": "Combined reason keys for overall status",
    "core_clustering_status": "Core clustering status",
    "source_diversity_status": "Source diversity status",
    "example_grounding_status": "Example grounding status",
    "overall_unknown_status": "All labeled rows unknown-rate status",
    "core_unknown_status": "Persona-core unknown-rate status",
    "core_coverage_status": "Persona-core coverage status",
    "effective_source_diversity_status": "Effective source-balance status",
    "source_concentration_status": "Labeled-source concentration status",
    "source_influence_concentration_status": "Blended source-influence concentration status",
    "weak_source_yield_status": "Weak-source cost-center status",
    "largest_cluster_dominance_status": "Largest cluster dominance status",
    "cluster_concentration_tail_status": "Top-3 cluster concentration status",
    "cluster_fragility_status": "Micro-cluster fragility status",
    "cluster_evidence_status": "Thin-evidence cluster status",
    "cluster_separation_status": "Minimum cluster separation status",
    "grounding_coverage_status": "Promoted persona grounding coverage status",
    "raw_record_rows": "Raw record row count (JSONL lines, not source count)",
    "normalized_post_rows": "Normalized post row count",
    "valid_candidate_rows": "Valid candidate post row count",
    "prefiltered_valid_rows": "Prefilter-retained valid post row count",
    "episode_rows": "Episode row count",
    "labeled_episode_rows": "Labeled episode row count",
    "persona_core_labeled_rows": "Persona-core labeled episode row count",
    "persona_core_coverage_of_all_labeled_pct": "Persona-core coverage of labeled episode rows (%)",
    "persona_core_unknown_ratio": "Persona-core unknown ratio",
    "overall_unknown_ratio": "All labeled rows unknown ratio",
    "effective_labeled_source_count": "Effective labeled-source count (source diversity score, not row count)",
    "effective_balanced_source_count": "Effective balanced-source count (influence-aware source balance score)",
    "largest_cluster_share_of_core_labeled": "Largest cluster share of persona-core labeled rows (%)",
    "top_3_cluster_share_of_core_labeled": "Top-3 cluster share of persona-core labeled rows",
    "robust_cluster_count": "Robust final cluster count",
    "stable_cluster_count": "Stable cluster count",
    "fragile_cluster_count": "Fragile cluster count",
    "micro_cluster_count": "Micro-cluster count",
    "thin_evidence_cluster_count": "Thin-evidence cluster count",
    "structurally_supported_cluster_count": "Structurally supported cluster count",
    "weak_separation_cluster_count": "Weak-separation cluster count",
    "fragile_tail_cluster_count": "Fragile tail cluster count",
    "fragile_tail_share_of_core_labeled": "Fragile tail share of persona-core labeled rows",
    "avg_cluster_separation": "Average cluster separation",
    "min_cluster_separation": "Minimum cluster separation",
    "largest_labeled_source_share_pct": "Largest labeled-source share of labeled episode rows (%)",
    "promoted_persona_episode_rows": "Promoted persona episode contribution rows",
    "grounded_promoted_persona_episode_rows": "Grounded promoted persona episode contribution rows",
    "largest_promoted_source_share_pct": "Largest promoted-persona source share (%)",
    "largest_grounded_source_share_pct": "Largest grounded-persona source share (%)",
    "largest_source_influence_share_pct": "Largest blended source influence share (%)",
    "weak_source_cost_center_count": "Weak-source cost-center count",
    "core_readiness_weak_source_cost_center_count": "Core-readiness weak-source cost-center count",
    "exploratory_only_weak_source_debt_count": "Exploratory-only weak-source debt count",
    "exploratory_only_weak_source_sources": "Exploratory-only weak-source debt source ids",
    "weak_source_denominator_policy_applied": "Weak-source denominator policy applied",
    "weak_source_denominator_policy_reason": "Weak-source denominator policy reason",
    "weak_source_cost_centers": "Weak-source cost-center source ids",
    "fix_now_source_count": "Fix-now source count",
    "tune_soon_source_count": "Tune-soon source count",
    "promoted_candidate_persona_count": "Promoted candidate persona count before grounding review",
    "promotion_visibility_persona_count": "Promotion-visibility persona count (review-visible promoted personas)",
    "headline_persona_count": "Headline persona count (final usable personas only)",
    "production_ready_persona_count": "Production-ready persona count (strict final usable personas only)",
    "review_ready_persona_count": "Review-ready persona count (visible for analyst review, not final usable)",
    "final_usable_persona_count": "Final usable persona count (structurally supported and grounded promoted personas only)",
    "deck_ready_persona_count": "Deck-ready persona count under current workbook gate",
    "promoted_persona_example_coverage_pct": "Promoted persona grounding coverage (%)",
    "promoted_persona_grounded_count": "Grounded promoted persona count",
    "promoted_persona_weakly_grounded_count": "Weakly grounded review-visible promoted persona count",
    "promoted_persona_ungrounded_count": "Ungrounded review-visible promoted persona count",
    "promoted_personas_weakly_grounded": "Weakly grounded promoted persona ids",
    "promoted_personas_missing_examples": "Promoted persona ids missing accepted grounding examples",
    "exploratory_bucket_count": "Exploratory non-promoted cluster count",
    "blocked_or_constrained_persona_count": "Blocked or constrained candidate persona count",
    "min_cluster_size": "Minimum cluster size for promotion review",
    "selected_axes": "Selected persona analysis axes",
    "clustering_mode": "Clustering mode",
}

QUALITY_CHECK_METRIC_LABELS = {
    "persona_readiness_state": "Persona readiness state",
    "persona_readiness_gate_status": "Persona readiness gate status",
    "persona_readiness_label": "Workbook interpretation label",
    "persona_asset_class": "Workbook asset class",
    "persona_readiness_summary": "Persona readiness summary",
    "persona_readiness_blockers": "Persona readiness blockers",
    "persona_usage_restriction": "Workbook usage restriction",
    "persona_completion_claim_allowed": "Persona-complete claim allowed",
    "promotion_constraint_status": "Promotion constraint status",
    "promotion_constraint_summary": "Promotion constraint summary",
    "persona_core_unknown_ratio": "Persona-core unknown ratio",
    "overall_unknown_ratio": "All labeled rows unknown ratio",
    "persona_core_coverage_of_all_labeled_pct": "Persona-core coverage of labeled episode rows (%)",
    "effective_labeled_source_count": "Effective labeled-source count (source diversity score, not row count)",
    "effective_balanced_source_count": "Effective balanced-source count (influence-aware source balance score)",
    "largest_labeled_source_share_pct": "Largest labeled-source share of labeled episode rows (%)",
    "promoted_persona_episode_rows": "Promoted persona episode contribution rows",
    "grounded_promoted_persona_episode_rows": "Grounded promoted persona episode contribution rows",
    "largest_source_influence_share_pct": "Largest blended source influence share (%)",
    "weak_source_cost_center_count": "Weak-source cost-center count",
    "core_readiness_weak_source_cost_center_count": "Core-readiness weak-source cost-center count",
    "exploratory_only_weak_source_debt_count": "Exploratory-only weak-source debt count",
    "exploratory_only_weak_source_sources": "Exploratory-only weak-source debt source ids",
    "weak_source_denominator_policy_applied": "Weak-source denominator policy applied",
    "weak_source_denominator_policy_reason": "Weak-source denominator policy reason",
    "largest_cluster_share_of_core_labeled": "Largest cluster share of persona-core labeled rows (%)",
    "top_3_cluster_share_of_core_labeled": "Top-3 cluster share of persona-core labeled rows",
    "micro_cluster_count": "Micro-cluster count",
    "thin_evidence_cluster_count": "Thin-evidence cluster count",
    "structurally_supported_cluster_count": "Structurally supported cluster count",
    "weak_separation_cluster_count": "Weak-separation cluster count",
    "fragile_tail_cluster_count": "Fragile tail cluster count",
    "fragile_tail_share_of_core_labeled": "Fragile tail share of persona-core labeled rows",
    "min_cluster_separation": "Minimum cluster separation",
    "promoted_persona_example_coverage_pct": "Promoted persona grounding coverage (%)",
    "promoted_persona_grounding_failure_count": "Promoted persona count failing grounded-usable policy",
    "selected_example_grounding_issue_count": "Selected representative example issue count (example rows)",
    "source_failures": "Sources with raw coverage but no labeled episode output",
    "quality_flag": "Quality flag derived from overall status",
    "overall_status": "Overall workbook status",
    "core_clustering_status": "Core clustering status",
    "source_diversity_status": "Source diversity status",
    "source_influence_concentration_status": "Blended source-influence concentration status",
    "weak_source_yield_status": "Weak-source cost-center status",
    "example_grounding_status": "Example grounding status",
    "denominator_consistency": "Denominator consistency contract status",
}

SOURCE_DISTRIBUTION_HEADER_OVERRIDES = {
    "raw_count": "raw_record_rows_for_source",
    "normalized_count": "normalized_post_rows_for_source",
    "valid_count": "valid_candidate_rows_for_source",
    "prefiltered_valid_count": "prefiltered_valid_post_rows_for_source",
    "episode_count": "episode_rows_for_source",
    "labeled_count": "labeled_episode_rows_for_source",
    "share_of_labeled": "share_of_labeled_episode_rows_pct",
}

COUNTS_HEADER_OVERRIDES = {
    "metric": "metric_key",
    "count": "row_count",
}

PERCENT_LIKE_COLUMNS = {
    "share_of_core_labeled",
    "share_of_all_labeled",
    "share_of_labeled",
    "pct_of_persona",
}

INTEGER_LIKE_COLUMNS = {
    "count",
    "persona_size",
    "denominator_value",
    "min_cluster_size",
    "example_rank",
    "mismatch_count",
    "critical_mismatch_count",
    "matched_axis_count",
    "grounded_candidate_count",
    "weak_candidate_count",
    "selected_example_count",
    "fallback_selected_count",
}


def export_workbook(
    root_dir: Path,
    overview_df: pd.DataFrame,
    counts_df: pd.DataFrame,
    source_distribution_df: pd.DataFrame,
    taxonomy_summary_df: pd.DataFrame,
    cluster_stats_df: pd.DataFrame,
    persona_summary_df: pd.DataFrame,
    persona_axes_df: pd.DataFrame,
    persona_needs_df: pd.DataFrame,
    persona_cooccurrence_df: pd.DataFrame,
    persona_examples_df: pd.DataFrame,
    quality_checks_df: pd.DataFrame,
    source_diagnostics_df: pd.DataFrame | None = None,
    quality_failures_df: pd.DataFrame | None = None,
    metric_glossary_df: pd.DataFrame | None = None,
) -> Path:
    """Write the final persona workbook from deterministic report tables."""
    return export_workbook_from_frames(
        root_dir=root_dir,
        frames={
            "overview": overview_df,
            "counts": counts_df,
            "source_distribution": source_distribution_df,
            "taxonomy_summary": taxonomy_summary_df,
            "cluster_stats": cluster_stats_df,
            "persona_summary": persona_summary_df,
            "persona_axes": persona_axes_df,
            "persona_needs": persona_needs_df,
            "persona_cooccurrence": persona_cooccurrence_df,
            "persona_examples": persona_examples_df,
            "quality_checks": quality_checks_df,
            "source_diagnostics": source_diagnostics_df if source_diagnostics_df is not None else pd.DataFrame(),
            "quality_failures": quality_failures_df if quality_failures_df is not None else pd.DataFrame(),
            "metric_glossary": metric_glossary_df if metric_glossary_df is not None else pd.DataFrame(),
        },
    )


def export_workbook_from_frames(root_dir: Path, frames: dict[str, pd.DataFrame]) -> Path:
    """Write the final workbook from the canonical workbook-frame mapping."""
    export_config = load_yaml(root_dir / "config" / "export_schema.yaml")
    workbook_name = export_config.get("workbook_name", "persona_pipeline_output.xlsx")
    output_path = ensure_dir(root_dir / "data" / "output") / workbook_name
    validated_frames = _prepare_workbook_frames(frames)
    sheet_specs = []
    for sheet_name in WORKBOOK_SHEET_NAMES:
        raw_frame = validated_frames.get(sheet_name, pd.DataFrame())
        export_frame = _display_frame(sheet_name, raw_frame)
        sheet_specs.append((sheet_name, raw_frame, export_frame))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, raw_frame, export_frame in sheet_specs:
            _prepare_for_excel(export_frame).to_excel(writer, sheet_name=sheet_name, index=False)
            _format_worksheet(writer.sheets[sheet_name], sheet_name, raw_frame, export_frame)
        _write_readme_sheet(writer.book)

    _verify_workbook_sheets(output_path)

    return output_path


def _prepare_workbook_frames(frames: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    """Validate and normalize workbook frames before export."""
    normalized = {sheet_name: round_frame_ratios(sheet_name, frames.get(sheet_name, pd.DataFrame())) for sheet_name in WORKBOOK_SHEET_NAMES}
    messages = validate_workbook_frames(normalized)
    for message in messages:
        warnings.warn(message, RuntimeWarning, stacklevel=2)
    if any(
        message.startswith("missing required sheet frame:")
        or message.startswith("sheet frame is null:")
        or message.startswith("forbidden generic share column:")
        or message.startswith("share denominator mismatch:")
        or message.startswith("legacy stage metric alias:")
        or message.startswith("stage metric mismatch:")
        or message.startswith("ambiguous persona count metric:")
        or message.startswith("missing persona promotion metric:")
        or message.startswith("persona promotion metric mismatch:")
        or message.startswith("ambiguous source_diagnostics column:")
        or message.startswith("missing source_diagnostics structure column:")
        or message.startswith("invalid source_diagnostics row placement:")
        or message.startswith("generic source_diagnostics reason:")
        or message.startswith("mixed-grain metric mislabeled as rate:")
        for message in messages
    ):
        missing = [message for message in messages if message.startswith("missing required sheet frame:") or message.startswith("sheet frame is null:")]
        denominator_errors = [
            message
            for message in messages
            if message.startswith("forbidden generic share column:")
            or message.startswith("share denominator mismatch:")
            or message.startswith("legacy stage metric alias:")
            or message.startswith("stage metric mismatch:")
            or message.startswith("ambiguous persona count metric:")
            or message.startswith("missing persona promotion metric:")
            or message.startswith("persona promotion metric mismatch:")
            or message.startswith("ambiguous source_diagnostics column:")
            or message.startswith("missing source_diagnostics structure column:")
            or message.startswith("invalid source_diagnostics row placement:")
            or message.startswith("generic source_diagnostics reason:")
            or message.startswith("mixed-grain metric mislabeled as rate:")
        ]
        failures = [*missing, *denominator_errors]
        raise ValueError("Workbook export validation failed: " + "; ".join(failures))
    return normalized


def _prepare_for_excel(df: pd.DataFrame, max_len: int = 32000) -> pd.DataFrame:
    """Trim long string cells for Excel export while keeping source artifacts untouched."""
    if df.empty:
        return df.copy()

    export_df = df.copy()
    for column in export_df.columns:
        export_df[column] = export_df[column].map(lambda value: _excel_safe_value(value, max_len=max_len))
    return export_df


def _display_frame(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """Return an export-only display frame with clearer reviewer-facing headers."""
    if df is None:
        return pd.DataFrame()
    frame = df.copy()
    if sheet_name == "overview":
        frame = _with_metric_display_label(frame, label_column="display_label", label_builder=lambda metric: _overview_metric_label(str(metric)))
        if "metric" in frame.columns:
            frame = frame.rename(columns={"metric": "metric_key", "value": "metric_value"})
        return frame
    if sheet_name == "quality_checks":
        frame = _with_metric_display_label(frame, label_column="display_label", label_builder=lambda metric: _quality_check_metric_label(str(metric)))
        rename_map = {
            "metric": "metric_key",
            "value": "metric_value",
            "threshold": "threshold_rule",
            "denominator_type": "denominator_type_key",
            "denominator_value": "denominator_row_count",
        }
        return frame.rename(columns={key: value for key, value in rename_map.items() if key in frame.columns})
    if sheet_name == "metric_glossary":
        frame = _with_metric_display_label(frame, label_column="workbook_label", label_builder=lambda metric: _glossary_metric_label(str(metric)))
        rename_map = {
            "metric": "metric_key",
            "denominator_type": "denominator_type_key",
        }
        return frame.rename(columns={key: value for key, value in rename_map.items() if key in frame.columns})
    if sheet_name == "counts":
        rename_map = {key: value for key, value in COUNTS_HEADER_OVERRIDES.items() if key in frame.columns}
        return frame.rename(columns=rename_map)
    if sheet_name == "source_distribution":
        rename_map = {key: value for key, value in SOURCE_DISTRIBUTION_HEADER_OVERRIDES.items() if key in frame.columns}
        return frame.rename(columns=rename_map)
    if sheet_name == "source_diagnostics":
        rename_map = {
            "grain": "row_grain",
            "metric_type": "metric_value_type",
            "denominator_value": "denominator_row_count",
        }
        return frame.rename(columns={key: value for key, value in rename_map.items() if key in frame.columns})
    if sheet_name in {"persona_summary", "cluster_stats"}:
        frame = _sort_persona_readiness_frame(frame)
        rename_map = {
            column: DISPLAY_HEADER_OVERRIDES.get(column, column)
            for column in frame.columns
        }
        frame = frame.rename(columns=rename_map)
        return _reorder_persona_display_frame(sheet_name, frame)
    rename_map = {
        column: DISPLAY_HEADER_OVERRIDES.get(column, column)
        for column in frame.columns
    }
    return frame.rename(columns=rename_map)


def _with_metric_display_label(df: pd.DataFrame, label_column: str, label_builder) -> pd.DataFrame:
    """Insert an export-only display label next to a metric key column."""
    if "metric" not in df.columns:
        return df
    frame = df.copy()
    insert_at = list(frame.columns).index("metric") + 1
    frame.insert(insert_at, label_column, frame["metric"].map(label_builder))
    return frame


def _overview_metric_label(metric: str) -> str:
    """Return a reviewer-facing label for overview metric keys."""
    return OVERVIEW_METRIC_LABELS.get(metric, metric.replace("_", " "))


def _quality_check_metric_label(metric: str) -> str:
    """Return a reviewer-facing label for quality check metric keys."""
    if ":" in metric:
        base_metric, source = metric.split(":", 1)
        source = source.strip()
        source_labels = {
            "valid_posts_per_normalized_post_pct": f"Source {source}: valid-post retention from normalized posts (%)",
            "prefiltered_valid_posts_per_valid_post_pct": f"Source {source}: relevance-prefilter retention from valid posts (%)",
            "labeled_episodes_per_episode_pct": f"Source {source}: labeled-episode retention from episode rows (%)",
            "labelable_episodes_per_labeled_episode_pct": f"Source {source}: labelable-episode retention from labeled episode rows (%)",
            "episodes_per_prefiltered_valid_post": f"Source {source}: episode rows per prefiltered valid post row",
            "labeled_episodes_per_prefiltered_valid_post": f"Source {source}: labeled episode rows per prefiltered valid post row",
            "labelable_episodes_per_prefiltered_valid_post": f"Source {source}: labelable episode rows per prefiltered valid post row",
        }
        return source_labels.get(base_metric, f"Source {source}: {base_metric.replace('_', ' ')}")
    return QUALITY_CHECK_METRIC_LABELS.get(metric, _overview_metric_label(metric))


def _glossary_metric_label(metric: str) -> str:
    """Return a reviewer-facing label for glossary rows."""
    return QUALITY_CHECK_METRIC_LABELS.get(metric, OVERVIEW_METRIC_LABELS.get(metric, metric.replace("_", " ")))


def _sort_persona_readiness_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Sort persona-facing sheets by readiness tier before existing share/size order."""
    if df.empty or "persona_id" not in df.columns:
        return df
    frame = df.copy()
    readiness_rank = {
        "production_ready_persona": 0,
        "review_ready_persona": 1,
        "blocked_or_constrained_candidate": 2,
        "exploratory_bucket": 3,
    }
    frame["_readiness_rank"] = frame.get("readiness_tier", pd.Series("", index=frame.index)).astype(str).map(readiness_rank).fillna(4).astype(int)
    frame["_production_rank"] = frame.get("production_ready_persona", pd.Series(False, index=frame.index)).fillna(False).astype(bool).astype(int)
    frame["_review_rank"] = frame.get("review_ready_persona", pd.Series(False, index=frame.index)).fillna(False).astype(bool).astype(int)
    share_column = None
    for candidate in ("share_of_core_labeled", "share_of_all_labeled"):
        if candidate in frame.columns:
            share_column = candidate
            frame[candidate] = pd.to_numeric(frame[candidate], errors="coerce")
            break
    if "persona_size" in frame.columns:
        frame["persona_size"] = pd.to_numeric(frame["persona_size"], errors="coerce")
    sort_columns = ["_readiness_rank", "_production_rank", "_review_rank"]
    ascending = [True, False, False]
    if share_column is not None:
        sort_columns.append(share_column)
        ascending.append(False)
    if "persona_size" in frame.columns:
        sort_columns.append("persona_size")
        ascending.append(False)
    sort_columns.append("persona_id")
    ascending.append(True)
    frame = frame.sort_values(sort_columns, ascending=ascending, kind="stable")
    return frame.drop(columns=["_readiness_rank", "_production_rank", "_review_rank"], errors="ignore")


def _reorder_persona_display_frame(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """Move readiness fields near the front for persona-facing workbook sheets."""
    preferred_prefix: list[str]
    if sheet_name == "persona_summary":
        preferred_prefix = [
            "persona_id",
            "persona_name",
            "persona_profile_name",
            "legacy_persona_name",
            "persona_size",
        ]
    else:
        preferred_prefix = [
            "persona_id",
            "persona_size",
        ]
    readiness_columns = [
        "readiness_tier",
        "production_ready_persona",
        "review_ready_persona",
        "review_visibility_status",
        "review_ready_reason",
        "blocked_reason",
        "workbook_policy_constraint",
    ]
    supporting_columns = [
        "workbook_readiness_state",
        "workbook_readiness_gate_status",
        "workbook_usage_restriction",
        "share_of_persona_core_labeled_pct",
        "share_of_all_labeled_pct",
    ]
    preferred = [column for column in [*preferred_prefix, *readiness_columns, *supporting_columns] if column in df.columns]
    remainder = [column for column in df.columns if column not in preferred]
    return df.loc[:, [*preferred, *remainder]]


def _format_worksheet(worksheet, sheet_name: str, raw_frame: pd.DataFrame, export_frame: pd.DataFrame) -> None:
    """Apply reviewer-focused formatting to one analytical sheet."""
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.sheet_view.showGridLines = True

    display_columns = list(export_frame.columns)
    for index, display_column in enumerate(display_columns, start=1):
        original_column = _raw_column_for_display_column(display_column)
        width = _column_width(display_column, worksheet, index)
        worksheet.column_dimensions[get_column_letter(index)].width = width
        _apply_column_format(worksheet, index, sheet_name, original_column, raw_frame)


def _raw_column_for_display_column(display_column: str) -> str:
    """Map an export-only display column back to the raw frame column when possible."""
    return DISPLAY_TO_RAW_HEADER_OVERRIDES.get(display_column, display_column)


def _column_width(display_column: str, worksheet, column_index: int) -> float:
    """Pick a practical Excel width for the column."""
    values = [len(str(display_column or ""))]
    for row_index in range(2, min(worksheet.max_row, 16) + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        values.append(len(str(value or "")))
    if any(token in str(display_column) for token in ["definition", "reason", "summary", "examples", "text", "why_"]):
        return 42
    if any(token in str(display_column) for token in ["metric", "status", "type", "grain", "source", "persona_id"]):
        return max(14, min(max(values) + 2, 28))
    return max(12, min(max(values) + 2, 22))


def _apply_column_format(worksheet, column_index: int, sheet_name: str, original_column: str, raw_frame: pd.DataFrame) -> None:
    """Apply numeric formatting based on semantic column meaning."""
    series = raw_frame.get(original_column, pd.Series(dtype=object)) if raw_frame is not None else pd.Series(dtype=object)
    is_numeric = not series.empty and pd.to_numeric(series, errors="coerce").notna().any()
    if not is_numeric:
        return
    number_format = None
    if original_column in PERCENT_LIKE_COLUMNS or str(original_column).endswith("_pct"):
        number_format = PERCENT_LITERAL_FORMAT
    elif original_column in INTEGER_LIKE_COLUMNS or str(original_column).endswith("_count") or str(original_column).endswith("_rows"):
        number_format = INTEGER_FORMAT
    elif str(original_column).endswith("_ratio") or str(original_column).endswith("_score") or str(original_column) in {"grounding_fit_score", "final_example_score", "metric_value"}:
        number_format = RATIO_FORMAT if str(original_column) in {"grounding_fit_score", "final_example_score", "metric_value"} else DECIMAL_FORMAT
    if sheet_name == "source_diagnostics" and original_column == "metric_value" and "metric_type" in raw_frame.columns:
        for row_index, metric_type in enumerate(raw_frame["metric_type"].astype(str).tolist(), start=2):
            cell = worksheet.cell(row=row_index, column=column_index)
            if metric_type == "percentage":
                cell.number_format = PERCENT_LITERAL_FORMAT
            elif metric_type == "count":
                cell.number_format = INTEGER_FORMAT
            elif metric_type == "ratio":
                cell.number_format = RATIO_FORMAT
            else:
                cell.number_format = "General"
        return
    if number_format is None:
        return
    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=column_index).number_format = number_format


def _write_readme_sheet(workbook) -> None:
    """Add a compact workbook guide and formula-backed provenance summary sheet."""
    if "readme" in workbook.sheetnames:
        del workbook["readme"]
    worksheet = workbook.create_sheet("readme", 0)
    rows = [
        ["Persona Workbook Guide", "Use this sheet first. Read the readiness gate before treating any persona sheet as final."],
        ["", ""],
        ["Readiness Gate", "Formula-backed links to the canonical readiness decision in overview."],
        ["Persona Readiness State", '=INDEX(overview!$C:$C,MATCH("persona_readiness_state",overview!$A:$A,0))'],
        ["Workbook Interpretation Label", '=INDEX(overview!$C:$C,MATCH("persona_readiness_label",overview!$A:$A,0))'],
        ["Workbook Asset Class", '=INDEX(overview!$C:$C,MATCH("persona_asset_class",overview!$A:$A,0))'],
        ["Persona Readiness Gate Status", '=INDEX(overview!$C:$C,MATCH("persona_readiness_gate_status",overview!$A:$A,0))'],
        ["Persona-Complete Claim Allowed", '=INDEX(overview!$C:$C,MATCH("persona_completion_claim_allowed",overview!$A:$A,0))'],
        ["Usage Restriction", '=INDEX(overview!$C:$C,MATCH("persona_usage_restriction",overview!$A:$A,0))'],
        ["Readiness Summary", '=INDEX(overview!$C:$C,MATCH("persona_readiness_summary",overview!$A:$A,0))'],
        ["Readiness Blockers", '=INDEX(overview!$C:$C,MATCH("persona_readiness_blockers",overview!$A:$A,0))'],
        ["Core-Readiness Weak-Source Cost-Center Count", '=INDEX(quality_checks!$C:$C,MATCH("core_readiness_weak_source_cost_center_count",quality_checks!$A:$A,0))'],
        ["Exploratory-Only Weak-Source Debt Count", '=INDEX(quality_checks!$C:$C,MATCH("exploratory_only_weak_source_debt_count",quality_checks!$A:$A,0))'],
        ["Exploratory-Only Weak-Source Debt Sources", '=INDEX(quality_checks!$C:$C,MATCH("exploratory_only_weak_source_sources",quality_checks!$A:$A,0))'],
        ["Weak-Source Denominator Policy Reason", '=INDEX(quality_checks!$C:$C,MATCH("weak_source_denominator_policy_reason",quality_checks!$A:$A,0))'],
        ["", ""],
        ["Pipeline Stage Summary", "Formula-backed links to the canonical stage metrics in overview."],
        ["Raw Record Row Count (JSONL lines, not sources)", '=INDEX(overview!$C:$C,MATCH("raw_record_rows",overview!$A:$A,0))'],
        ["Normalized Post Row Count", '=INDEX(overview!$C:$C,MATCH("normalized_post_rows",overview!$A:$A,0))'],
        ["Valid Candidate Post Row Count", '=INDEX(overview!$C:$C,MATCH("valid_candidate_rows",overview!$A:$A,0))'],
        ["Prefilter-Retained Valid Post Row Count", '=INDEX(overview!$C:$C,MATCH("prefiltered_valid_rows",overview!$A:$A,0))'],
        ["Episode Row Count", '=INDEX(overview!$C:$C,MATCH("episode_rows",overview!$A:$A,0))'],
        ["Labeled Episode Row Count", '=INDEX(overview!$C:$C,MATCH("labeled_episode_rows",overview!$A:$A,0))'],
        ["Persona-Core Labeled Episode Row Count", '=INDEX(overview!$C:$C,MATCH("persona_core_labeled_rows",overview!$A:$A,0))'],
        ["Review-Visible Promoted Persona Count", '=INDEX(overview!$C:$C,MATCH("promotion_visibility_persona_count",overview!$A:$A,0))'],
        ["Headline Persona Count (final usable personas only)", '=INDEX(overview!$C:$C,MATCH("headline_persona_count",overview!$A:$A,0))'],
        ["Production-Ready Persona Count", '=INDEX(overview!$C:$C,MATCH("production_ready_persona_count",overview!$A:$A,0))'],
        ["Review-Ready Persona Count", '=INDEX(overview!$C:$C,MATCH("review_ready_persona_count",overview!$A:$A,0))'],
        ["Final Usable Persona Count (structurally supported and grounded promoted personas only)", '=INDEX(overview!$C:$C,MATCH("final_usable_persona_count",overview!$A:$A,0))'],
        ["Deck-Ready Persona Count", '=INDEX(overview!$C:$C,MATCH("deck_ready_persona_count",overview!$A:$A,0))'],
        ["Blocked or Constrained Persona Count", '=INDEX(overview!$C:$C,MATCH("blocked_or_constrained_persona_count",overview!$A:$A,0))'],
        ["Exploratory Bucket Count", '=INDEX(overview!$C:$C,MATCH("exploratory_bucket_count",overview!$A:$A,0))'],
        ["Weakly Grounded Review-Visible Promoted Persona Count", '=INDEX(overview!$C:$C,MATCH("promoted_persona_weakly_grounded_count",overview!$A:$A,0))'],
        ["Ungrounded Review-Visible Promoted Persona Count", '=INDEX(overview!$C:$C,MATCH("promoted_persona_ungrounded_count",overview!$A:$A,0))'],
        ["Approximate Unknown Labeled Episode Rows", '=ROUND(INDEX(overview!$C:$C,MATCH("overall_unknown_ratio",overview!$A:$A,0))*INDEX(overview!$C:$C,MATCH("labeled_episode_rows",overview!$A:$A,0)),0)'],
        ["", ""],
        ["How To Read Denominators", ""],
        ["share_of_persona_core_labeled_pct", "Percentage over persona_core_labeled_rows. This denominator is persona-core labeled episode rows, not all labeled rows."],
        ["share_of_all_labeled_pct", "Percentage over all labeled_episode_rows. Use this only for whole-workbook context, not persona-core clustering coverage."],
        ["row_grain", "The entity counted by the row: post rows, episode rows, mixed_grain_bridge ratios, or other non-funnel metrics."],
        ["denominator_type_key", "The semantic denominator family. Cross-check this against metric_glossary before comparing row counts, source counts, or persona counts."],
        ["", ""],
        ["Review Tips", ""],
        ["Readiness gate", "If persona_readiness_state is below deck_ready, no sheet in this workbook may be interpreted as a final persona asset. Treat persona_summary and cluster_stats as hypothesis or review material only."],
        ["Production-ready personas", "Production-ready personas are strict final usable outputs."],
        ["Review-ready personas", "Review-ready personas are strong candidates for analyst review, but are not included in final usable persona count."],
        ["Threshold discipline", "Review-ready status does not relax workbook policy or production-ready thresholds."],
        ["Human review requirement", "Review-ready personas need human review before deck-ready or production use."],
        ["Weak-source split", "Weak-source Cost-Center Count remains the full visible diagnostic count. Core-Readiness Weak-Source Cost-Center Count is the subset still used for workbook hard-fail pressure. Exploratory-Only Weak-Source Debt stays visible but is separated from core reviewability pressure."],
        ["Grounding states", "See persona_summary and cluster_stats for base_promotion_status, structural_support_status, visibility_state, usability_state, deck_readiness_state, promotion_action, promotion_status, grounding_status, promotion_grounding_status, and reporting_readiness_status. Review-visible personas remain workbook-visible for audit but are not final usable or deck-ready personas."],
        ["Persona counts", "Use Headline Persona Count or Final Usable Persona Count for headline or downstream persona totals only when persona_readiness_state is deck_ready or higher. Production-Ready Persona Count mirrors final usable personas. Review-Ready Persona Count is reported separately and must not be added into final usable totals."],
        ["Rows versus sources", "Raw Record Row Count is a count of JSONL rows. Effective labeled-source count and source_distribution rows describe sources, not post or episode rows."],
        ["Mixed-grain diagnostics", "source_diagnostics rows with row_grain=mixed_grain_bridge are cross-grain ratios, not same-grain funnel percentages."],
        ["Glossary", "See metric_glossary for metric keys, reviewer-facing workbook labels, and denominator semantics."],
    ]
    for row in rows:
        worksheet.append(row)
    worksheet.freeze_panes = "A3"
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 92
    worksheet["A1"].font = TITLE_FONT
    worksheet["A3"].font = HEADER_FONT
    worksheet["B3"].font = HEADER_FONT
    worksheet["A13"].font = HEADER_FONT
    worksheet["B13"].font = HEADER_FONT
    for cell in worksheet[3]:
        cell.fill = HEADER_FILL
    for cell in worksheet[13]:
        cell.fill = HEADER_FILL
    for row_index in [37, 43]:
        for cell in worksheet[row_index]:
            cell.font = HEADER_FONT
            cell.fill = SUBTLE_FILL
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    glossary_row = len(rows)
    worksheet[f"B{glossary_row}"].hyperlink = "#metric_glossary!A1"
    worksheet[f"B{glossary_row}"].style = "Hyperlink"


def _excel_safe_value(value: object, max_len: int) -> object:
    """Convert nested values to strings and enforce Excel cell length limits."""
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str):
        value = ILLEGAL_EXCEL_CHAR_RE.sub("", value)
    if isinstance(value, str) and len(value) > max_len:
        return value[:max_len] + "...[truncated]"
    return value


def _verify_workbook_sheets(path: Path) -> None:
    """Verify the written workbook contains the required sheets."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    missing = [sheet_name for sheet_name in WORKBOOK_SHEET_NAMES if sheet_name not in sheet_names]
    if missing:
        raise ValueError(f"Workbook missing required sheets: {', '.join(missing)}")
