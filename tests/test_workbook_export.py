"""Tests for canonical workbook bundle and xlsx export."""

from __future__ import annotations

import tempfile
import unittest
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.analysis.workbook_bundle import assemble_workbook_frames, validate_workbook_frames
from src.exporters.xlsx_exporter import export_workbook_from_frames


class WorkbookExportTests(unittest.TestCase):
    """Verify workbook creation remains robust for sparse inputs."""

    @staticmethod
    def _readme_value_by_label(workbook, label: str) -> object:
        for left_cell, right_cell in workbook["readme"].iter_rows(min_col=1, max_col=2, values_only=True):
            if left_cell == label:
                return right_cell
        raise AssertionError(f"Missing readme label: {label}")

    @staticmethod
    def _sheet_rows_by_first_column(worksheet) -> dict[object, tuple[object, ...]]:
        rows = list(worksheet.iter_rows(values_only=True))
        return {row[0]: row for row in rows[1:] if row and row[0] is not None}

    def test_export_creates_required_sheets_even_with_sparse_frames(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            empty = pd.DataFrame()
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame({"metric": ["x"], "value": ["y"]}),
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [1]}),
                source_distribution_df=empty,
                taxonomy_summary_df=empty,
                cluster_stats_df=empty,
                persona_summary_df=empty,
                persona_axes_df=empty,
                persona_needs_df=empty,
                persona_cooccurrence_df=empty,
                persona_examples_df=empty,
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"], "threshold": [""], "status": ["pass"], "notes": [""]}),
            )
            output = export_workbook_from_frames(root, frames)
            workbook = load_workbook(output, read_only=False)
            try:
                self.assertEqual(workbook.sheetnames[0], "readme")
                self.assertEqual(workbook.sheetnames[1:], [
                    "overview",
                    "counts",
                    "source_distribution",
                    "taxonomy_summary",
                    "cluster_stats",
                    "persona_summary",
                    "persona_axes",
                    "persona_needs",
                    "persona_cooccurrence",
                    "persona_examples",
                    "quality_checks",
                    "source_diagnostics",
                    "quality_failures",
                    "metric_glossary",
                ])
                self.assertEqual(workbook["overview"].freeze_panes, "A2")
                self.assertIsNotNone(workbook["overview"].auto_filter.ref)
                self.assertEqual(str(workbook["readme"]["B4"].value).startswith("=INDEX("), True)
                self.assertEqual(str(self._readme_value_by_label(workbook, "Core-Readiness Weak-Source Cost-Center Count")).startswith("=INDEX("), True)
                self.assertIn(
                    "Exploratory-Only Weak-Source Debt",
                    str(self._readme_value_by_label(workbook, "Weak-source split")),
                )
                overview_headers = [cell.value for cell in next(workbook["overview"].iter_rows(min_row=1, max_row=1))]
                self.assertEqual(overview_headers[:3], ["metric_key", "display_label", "metric_value"])
            finally:
                workbook.close()

    def test_validate_workbook_frames_warns_on_sparse_optional_sheets(self) -> None:
        frames = assemble_workbook_frames(
            overview_df=pd.DataFrame({"metric": ["x"], "value": ["y"]}),
            counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [1]}),
            source_distribution_df=pd.DataFrame({"source": ["reddit"], "normalized_count": [1], "valid_count": [1], "episode_count": [1], "labeled_count": [1], "share_of_labeled": [100.0]}),
            taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"], "why_it_matters": ["x"], "allowed_values_or_logic": ["a"], "evidence_fields": ["b"]}),
            cluster_stats_df=pd.DataFrame(),
            persona_summary_df=pd.DataFrame(),
            persona_axes_df=pd.DataFrame(),
            persona_needs_df=pd.DataFrame(),
            persona_cooccurrence_df=pd.DataFrame(),
            persona_examples_df=pd.DataFrame(),
            quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["LOW QUALITY"], "threshold": [""], "status": ["warn"], "notes": ["x"]}),
        )
        messages = validate_workbook_frames(frames)
        self.assertTrue(any("sparse data: empty cluster_stats sheet" == msg for msg in messages))

    def test_validate_workbook_frames_warns_on_missing_optional_columns(self) -> None:
        frames = {
            "overview": pd.DataFrame({"metric": ["x"]}),
            "counts": pd.DataFrame({"metric": ["raw_record_rows"], "count": [1]}),
            "source_distribution": pd.DataFrame({"source": ["reddit"]}),
            "taxonomy_summary": pd.DataFrame({"axis_name": ["role"]}),
            "cluster_stats": pd.DataFrame(),
            "persona_summary": pd.DataFrame(),
            "persona_axes": pd.DataFrame(),
            "persona_needs": pd.DataFrame(),
            "persona_cooccurrence": pd.DataFrame(),
            "persona_examples": pd.DataFrame(),
            "quality_checks": pd.DataFrame({"metric": ["quality_flag"]}),
            "source_diagnostics": pd.DataFrame(),
            "quality_failures": pd.DataFrame(),
            "metric_glossary": pd.DataFrame(),
        }
        messages = validate_workbook_frames(frames)
        self.assertTrue(any(msg == "missing optional column: overview.value" for msg in messages))

    def test_export_warns_for_invalid_ratio_but_still_writes_sparse_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame(
                    {
                        "metric": [
                            "x",
                            "promoted_candidate_persona_count",
                            "promotion_visibility_persona_count",
                            "headline_persona_count",
                            "production_ready_persona_count",
                            "review_ready_persona_count",
                            "final_usable_persona_count",
                            "exploratory_bucket_count",
                            "deck_ready_persona_count",
                        ],
                        "value": ["y", 1, 1, 1, 1, 0, 1, 0, 1],
                    }
                ),
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [1]}),
                source_distribution_df=pd.DataFrame({"source": ["reddit"], "normalized_count": [1], "valid_count": [1], "episode_count": [1], "labeled_count": [1], "share_of_labeled": [100.0]}),
                taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"], "why_it_matters": ["x"], "allowed_values_or_logic": ["a"], "evidence_fields": ["b"]}),
                cluster_stats_df=pd.DataFrame(
                    {
                        "persona_id": ["p1"],
                        "persona_size": [1],
                        "share_of_core_labeled": ["bad"],
                        "share_of_all_labeled": [100.0],
                        "base_promotion_status": ["promoted_candidate_persona"],
                        "promotion_status": ["promoted_persona"],
                        "promotion_grounding_status": ["promoted_and_grounded"],
                        "final_usable_persona": [True],
                        "denominator_type": ["persona_core_labeled_rows"],
                        "denominator_value": [1],
                        "dominant_signature": [""],
                        "dominant_bottleneck": [""],
                        "dominant_analysis_goal": [""],
                    }
                ),
                persona_summary_df=pd.DataFrame(),
                persona_axes_df=pd.DataFrame(),
                persona_needs_df=pd.DataFrame(),
                persona_cooccurrence_df=pd.DataFrame(),
                persona_examples_df=pd.DataFrame(),
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"], "threshold": [""], "status": ["pass"], "notes": [""]}),
            )
            with warnings.catch_warnings(record=True) as captured:
                warnings.simplefilter("always")
                output = export_workbook_from_frames(root, frames)
            self.assertTrue(output.exists())
            self.assertTrue(any("non-numeric ratio values: cluster_stats.share_of_core_labeled" in str(item.message) for item in captured))

    def test_export_applies_literal_percentage_format_to_pct_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame(
                    {
                        "metric": [
                            "quality_flag",
                            "promoted_candidate_persona_count",
                            "promotion_visibility_persona_count",
                            "headline_persona_count",
                            "production_ready_persona_count",
                            "review_ready_persona_count",
                            "final_usable_persona_count",
                            "exploratory_bucket_count",
                            "deck_ready_persona_count",
                        ],
                        "value": ["OK", 1, 1, 1, 1, 0, 1, 0, 1],
                    }
                ),
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [1]}),
                source_distribution_df=pd.DataFrame({"source": ["reddit"], "raw_count": [1], "normalized_count": [1], "valid_count": [1], "prefiltered_valid_count": [1], "episode_count": [1], "labeled_count": [1], "share_of_labeled": [100.0]}),
                taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"], "why_it_matters": ["x"], "allowed_values_or_logic": ["a"], "evidence_fields": ["b"]}),
                cluster_stats_df=pd.DataFrame({"persona_id": ["p1"], "persona_size": [1], "share_of_core_labeled": [75.0], "share_of_all_labeled": [50.0], "base_promotion_status": ["promoted_candidate_persona"], "promotion_status": ["promoted_persona"], "promotion_grounding_status": ["promoted_and_grounded"], "final_usable_persona": [True], "denominator_type": ["persona_core_labeled_rows"], "denominator_value": [1], "dominant_signature": [""], "dominant_bottleneck": [""], "dominant_analysis_goal": [""]}),
                persona_summary_df=pd.DataFrame(),
                persona_axes_df=pd.DataFrame(),
                persona_needs_df=pd.DataFrame(),
                persona_cooccurrence_df=pd.DataFrame(),
                persona_examples_df=pd.DataFrame(),
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"], "threshold": [""], "status": ["pass"], "notes": [""]}),
            )
            output = export_workbook_from_frames(root, frames)
            workbook = load_workbook(output, read_only=False)
            try:
                worksheet = workbook["cluster_stats"]
                self.assertEqual(worksheet["C2"].number_format, '0.0"%"')
                self.assertEqual(worksheet["D2"].number_format, '0.0"%"')
            finally:
                workbook.close()

    def test_export_fails_when_share_name_and_denominator_do_not_match(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame({"metric": ["x", "promoted_candidate_persona_count", "promotion_visibility_persona_count", "headline_persona_count", "final_usable_persona_count", "deck_ready_persona_count"], "value": ["y", 1, 1, 1, 1, 1]}),
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [1]}),
                source_distribution_df=pd.DataFrame({"source": ["reddit"], "normalized_count": [1], "valid_count": [1], "episode_count": [1], "labeled_count": [1], "share_of_labeled": [100.0]}),
                taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"], "why_it_matters": ["x"], "allowed_values_or_logic": ["a"], "evidence_fields": ["b"]}),
                cluster_stats_df=pd.DataFrame(
                    {
                        "persona_id": ["p1"],
                        "persona_size": [1],
                        "share_of_core_labeled": [100.0],
                        "base_promotion_status": ["promoted_candidate_persona"],
                        "promotion_status": ["promoted_persona"],
                        "promotion_grounding_status": ["promoted_and_grounded"],
                        "final_usable_persona": [True],
                        "denominator_type": ["labeled_episode_rows"],
                        "denominator_value": [1],
                        "dominant_signature": [""],
                        "dominant_bottleneck": [""],
                        "dominant_analysis_goal": [""],
                    }
                ),
                persona_summary_df=pd.DataFrame(),
                persona_axes_df=pd.DataFrame(),
                persona_needs_df=pd.DataFrame(),
                persona_cooccurrence_df=pd.DataFrame(),
                persona_examples_df=pd.DataFrame(),
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"], "threshold": [""], "status": ["pass"], "notes": [""]}),
            )
            with self.assertRaisesRegex(ValueError, "share denominator mismatch"):
                export_workbook_from_frames(root, frames)

    def test_export_adds_explicit_labels_and_keeps_glossary_definitions_aligned(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame(
                    {
                        "metric": [
                            "raw_record_rows",
                            "promotion_visibility_persona_count",
                            "headline_persona_count",
                            "production_ready_persona_count",
                            "review_ready_persona_count",
                            "final_usable_persona_count",
                            "blocked_or_constrained_persona_count",
                        ],
                        "value": [12, 3, 2, 2, 1, 2, 1],
                    }
                ),
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [12]}),
                source_distribution_df=pd.DataFrame(
                    {
                        "source": ["reddit"],
                        "raw_count": [12],
                        "normalized_count": [10],
                        "valid_count": [8],
                        "prefiltered_valid_count": [4],
                        "episode_count": [5],
                        "labeled_count": [5],
                        "share_of_labeled": [100.0],
                        "denominator_type": ["labeled_episode_rows"],
                        "denominator_value": [5],
                    }
                ),
                taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"], "why_it_matters": ["x"], "allowed_values_or_logic": ["a"], "evidence_fields": ["b"]}),
                cluster_stats_df=pd.DataFrame(),
                persona_summary_df=pd.DataFrame(),
                persona_axes_df=pd.DataFrame(),
                persona_needs_df=pd.DataFrame(),
                persona_cooccurrence_df=pd.DataFrame(),
                persona_examples_df=pd.DataFrame(),
                quality_checks_df=pd.DataFrame(
                    {
                        "metric": ["effective_labeled_source_count", "promoted_persona_grounding_failure_count"],
                        "value": [3.2, 1],
                        "threshold": ["fail<4.0", "warn<100.0; fail<80.0"],
                        "status": ["fail", "fail"],
                        "level": ["hard_fail", "hard_fail"],
                        "denominator_type": ["", "promoted_persona_rows"],
                        "denominator_value": ["", 3],
                        "notes": ["effective_source_diversity_low", "promoted_persona_examples_missing"],
                    }
                ),
                source_diagnostics_df=pd.DataFrame(),
                quality_failures_df=pd.DataFrame(),
                metric_glossary_df=pd.DataFrame(
                    {
                        "metric": ["effective_labeled_source_count", "final_usable_persona_count", "promotion_visibility_persona_count"],
                        "denominator_type": ["effective_labeled_source_count", "persona_cluster_rows", "persona_cluster_rows"],
                        "definition": [
                            "Effective count of contributing labeled sources after fractional down-weighting for very small labeled-source volumes. This is a source-count metric, not a row-count metric.",
                            "Count of final usable personas for downstream reporting. Under the current policy this includes only promoted_and_grounded personas, not weakly grounded or ungrounded review-visible personas.",
                            "Count of promoted personas that remain visible in the workbook for reviewer inspection after grounding policy merge. Under the current flag policy this includes grounded, weakly grounded, and ungrounded promoted personas.",
                        ],
                    }
                ),
            )
            output = export_workbook_from_frames(root, frames)
            workbook = load_workbook(output, read_only=True)
            try:
                overview_rows = self._sheet_rows_by_first_column(workbook["overview"])
                quality_rows = list(workbook["quality_checks"].iter_rows(min_row=1, max_row=3, values_only=True))
                glossary_rows = list(workbook["metric_glossary"].iter_rows(min_row=1, max_row=4, values_only=True))
                readme_persona_copy = self._readme_value_by_label(workbook, "Persona counts")
                readme_row_source_copy = self._readme_value_by_label(workbook, "Rows versus sources")
            finally:
                workbook.close()

            self.assertEqual(overview_rows["raw_record_rows"][:3], ("raw_record_rows", "Raw record row count (JSONL lines, not source count)", 12))
            self.assertEqual(overview_rows["promotion_visibility_persona_count"][:3], ("promotion_visibility_persona_count", "Promotion-visibility persona count (review-visible promoted personas)", 3))
            self.assertEqual(overview_rows["headline_persona_count"][:3], ("headline_persona_count", "Headline persona count (final usable personas only)", 2))
            self.assertEqual(overview_rows["production_ready_persona_count"][:3], ("production_ready_persona_count", "Production-ready persona count (strict final usable personas only)", 2))
            self.assertEqual(overview_rows["review_ready_persona_count"][:3], ("review_ready_persona_count", "Review-ready persona count (visible for analyst review, not final usable)", 1))
            self.assertEqual(overview_rows["final_usable_persona_count"][:3], ("final_usable_persona_count", "Final usable persona count (structurally supported and grounded promoted personas only)", 2))
            self.assertEqual(overview_rows["blocked_or_constrained_persona_count"][:3], ("blocked_or_constrained_persona_count", "Blocked or constrained candidate persona count", 1))

            self.assertEqual(quality_rows[0][:3], ("metric_key", "display_label", "metric_value"))
            self.assertEqual(quality_rows[1][1], "Effective labeled-source count (source diversity score, not row count)")
            self.assertEqual(quality_rows[2][1], "Promoted persona count failing grounded-usable policy")

            self.assertEqual(glossary_rows[0][:3], ("metric_key", "workbook_label", "denominator_type_key"))
            self.assertEqual(glossary_rows[1][1], "Effective labeled-source count (source diversity score, not row count)")
            self.assertIn("source-count metric, not a row-count metric", str(glossary_rows[1][3]))
            self.assertIn("Review-Ready Persona Count is reported separately", str(readme_persona_copy))
            self.assertIn("Raw Record Row Count is a count of JSONL rows", str(readme_row_source_copy))

    def test_validate_workbook_frames_rejects_final_asset_claim_below_deck_ready(self) -> None:
        frames = assemble_workbook_frames(
            overview_df=pd.DataFrame(
                {
                    "metric": [
                        "persona_readiness_state",
                        "persona_asset_class",
                        "persona_completion_claim_allowed",
                        "promoted_candidate_persona_count",
                        "promotion_visibility_persona_count",
                        "headline_persona_count",
                        "final_usable_persona_count",
                        "deck_ready_persona_count",
                    ],
                    "value": ["exploratory_only", "final_persona_asset", True, 1, 1, 1, 1, 0],
                }
            ),
            counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [12]}),
            source_distribution_df=pd.DataFrame(),
            taxonomy_summary_df=pd.DataFrame(),
            cluster_stats_df=pd.DataFrame(
                {
                    "persona_id": ["persona_01"],
                    "persona_size": [1],
                    "share_of_core_labeled": [100.0],
                    "share_of_all_labeled": [100.0],
                    "base_promotion_status": ["promoted_candidate_persona"],
                    "promotion_status": ["promoted_persona"],
                    "workbook_review_visible": [True],
                    "promotion_grounding_status": ["promoted_and_grounded"],
                    "final_usable_persona": [True],
                    "denominator_type": ["persona_core_labeled_rows"],
                    "denominator_value": [1],
                    "dominant_signature": ["workflow_stage=reporting"],
                    "dominant_bottleneck": ["manual_reporting"],
                    "dominant_analysis_goal": ["report_speed"],
                }
            ),
            persona_summary_df=pd.DataFrame(),
            persona_axes_df=pd.DataFrame(),
            persona_needs_df=pd.DataFrame(),
            persona_cooccurrence_df=pd.DataFrame(),
            persona_examples_df=pd.DataFrame(),
            quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["UNSTABLE"], "threshold": [""], "status": ["fail"], "level": ["hard_fail"], "denominator_type": [""], "denominator_value": [""], "notes": [""]}),
            source_diagnostics_df=pd.DataFrame(),
            quality_failures_df=pd.DataFrame(),
            metric_glossary_df=pd.DataFrame(),
        )

        messages = validate_workbook_frames(frames)
        self.assertIn("persona readiness metric mismatch: final persona asset class is forbidden below deck_ready", messages)

    def test_validate_workbook_frames_rejects_persona_sheet_readiness_drift(self) -> None:
        frames = assemble_workbook_frames(
            overview_df=pd.DataFrame(
                {
                    "metric": [
                        "persona_readiness_state",
                        "persona_readiness_gate_status",
                        "persona_asset_class",
                        "persona_completion_claim_allowed",
                        "persona_usage_restriction",
                        "promoted_candidate_persona_count",
                        "promotion_visibility_persona_count",
                        "headline_persona_count",
                        "final_usable_persona_count",
                        "deck_ready_persona_count",
                    ],
                    "value": ["exploratory_only", "FAIL", "hypothesis_material", False, "Hypothesis material only. Not a final persona asset.", 1, 1, 1, 1, 0],
                }
            ),
            counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [12]}),
            source_distribution_df=pd.DataFrame(),
            taxonomy_summary_df=pd.DataFrame(),
            cluster_stats_df=pd.DataFrame(
                {
                    "persona_id": ["persona_01"],
                    "workbook_readiness_state": ["deck_ready"],
                    "workbook_readiness_gate_status": ["OK"],
                    "workbook_usage_restriction": ["Final persona asset."],
                    "persona_size": [1],
                    "share_of_core_labeled": [100.0],
                    "share_of_all_labeled": [100.0],
                    "base_promotion_status": ["promoted_candidate_persona"],
                    "promotion_status": ["promoted_persona"],
                    "workbook_review_visible": [True],
                    "promotion_grounding_status": ["promoted_and_grounded"],
                    "final_usable_persona": [True],
                    "denominator_type": ["persona_core_labeled_rows"],
                    "denominator_value": [1],
                    "dominant_signature": ["workflow_stage=reporting"],
                    "dominant_bottleneck": ["manual_reporting"],
                    "dominant_analysis_goal": ["report_speed"],
                }
            ),
            persona_summary_df=pd.DataFrame(
                {
                    "persona_id": ["persona_01"],
                    "workbook_readiness_state": ["exploratory_only"],
                    "workbook_readiness_gate_status": ["FAIL"],
                    "workbook_usage_restriction": ["Hypothesis material only. Not a final persona asset."],
                }
            ),
            persona_axes_df=pd.DataFrame(),
            persona_needs_df=pd.DataFrame(),
            persona_cooccurrence_df=pd.DataFrame(),
            persona_examples_df=pd.DataFrame(),
            quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["UNSTABLE"], "threshold": [""], "status": ["fail"], "level": ["hard_fail"], "denominator_type": [""], "denominator_value": [""], "notes": [""]}),
            source_diagnostics_df=pd.DataFrame(),
            quality_failures_df=pd.DataFrame(),
            metric_glossary_df=pd.DataFrame(),
        )

        messages = validate_workbook_frames(frames)
        self.assertIn("persona readiness metric mismatch: cluster_stats.workbook_readiness_state must match overview.persona_readiness_state", messages)
        self.assertIn("persona readiness metric mismatch: cluster_stats.workbook_readiness_gate_status must match overview.persona_readiness_gate_status", messages)
        self.assertIn("persona readiness metric mismatch: cluster_stats.workbook_usage_restriction must match overview.persona_usage_restriction", messages)

    def test_export_readme_blocks_finality_below_deck_ready(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame(
                    {
                        "metric": [
                            "persona_readiness_state",
                            "persona_readiness_label",
                            "persona_asset_class",
                            "persona_readiness_gate_status",
                            "persona_completion_claim_allowed",
                            "persona_usage_restriction",
                            "persona_readiness_summary",
                            "persona_readiness_blockers",
                            "promotion_visibility_persona_count",
                            "headline_persona_count",
                            "final_usable_persona_count",
                            "deck_ready_persona_count",
                            "promoted_persona_weakly_grounded_count",
                            "promoted_persona_ungrounded_count",
                            "overall_unknown_ratio",
                            "labeled_episode_rows",
                        ],
                        "value": [
                            "exploratory_only",
                            "Hypothesis Material",
                            "hypothesis_material",
                            "FAIL",
                            False,
                            "Hypothesis material only. Not a final persona asset.",
                            "Exploratory workbook only.",
                            "overall_unknown_ratio<=0.30",
                            3,
                            0,
                            2,
                            2,
                            0,
                            1,
                            0.388,
                            100,
                        ],
                    }
                ),
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [12]}),
                source_distribution_df=pd.DataFrame(),
                taxonomy_summary_df=pd.DataFrame(),
                cluster_stats_df=pd.DataFrame(),
                persona_summary_df=pd.DataFrame(),
                persona_axes_df=pd.DataFrame(),
                persona_needs_df=pd.DataFrame(),
                persona_cooccurrence_df=pd.DataFrame(),
                persona_examples_df=pd.DataFrame(),
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["UNSTABLE"], "threshold": [""], "status": ["fail"], "level": ["hard_fail"], "denominator_type": [""], "denominator_value": [""], "notes": [""]}),
                source_diagnostics_df=pd.DataFrame(),
                quality_failures_df=pd.DataFrame(),
                metric_glossary_df=pd.DataFrame(),
            )

            output = export_workbook_from_frames(root, frames)
            workbook = load_workbook(output, read_only=True)
            try:
                readme_gate_copy = self._readme_value_by_label(workbook, "Readiness gate")
                readme_persona_count_copy = self._readme_value_by_label(workbook, "Persona counts")
            finally:
                workbook.close()

            self.assertIn("no sheet in this workbook may be interpreted as a final persona asset", str(readme_gate_copy))
            self.assertIn("only when persona_readiness_state is deck_ready or higher", str(readme_persona_count_copy))

    def test_export_surfaces_review_ready_tier_without_changing_headline_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            overview_df = pd.DataFrame(
                {
                    "metric": [
                        "persona_readiness_state",
                        "persona_readiness_label",
                        "persona_asset_class",
                        "persona_readiness_gate_status",
                        "persona_completion_claim_allowed",
                        "persona_usage_restriction",
                        "persona_readiness_summary",
                        "persona_readiness_blockers",
                        "raw_record_rows",
                        "normalized_post_rows",
                        "valid_candidate_rows",
                        "prefiltered_valid_rows",
                        "episode_rows",
                        "labeled_episode_rows",
                        "persona_core_labeled_rows",
                        "promoted_candidate_persona_count",
                        "promotion_visibility_persona_count",
                        "headline_persona_count",
                        "production_ready_persona_count",
                        "review_ready_persona_count",
                        "final_usable_persona_count",
                        "deck_ready_persona_count",
                        "blocked_or_constrained_persona_count",
                        "exploratory_bucket_count",
                        "promoted_persona_weakly_grounded_count",
                        "promoted_persona_ungrounded_count",
                        "overall_unknown_ratio",
                    ],
                    "value": [
                        "reviewable_but_not_deck_ready",
                        "Reviewable Candidate",
                        "review_only_material",
                        "WARN",
                        False,
                        "Review material only. Not a final persona asset.",
                        "Workbook is reviewable but not deck-ready.",
                        "top_3_cluster_share_of_core_labeled",
                        100,
                        90,
                        80,
                        70,
                        60,
                        60,
                        55,
                        5,
                        4,
                        3,
                        3,
                        1,
                        3,
                        0,
                        1,
                        0,
                        0,
                        0,
                        0.14,
                    ],
                }
            )
            cluster_stats_df = pd.DataFrame(
                {
                    "persona_id": ["persona_01", "persona_02", "persona_03", "persona_04", "persona_05"],
                    "workbook_readiness_state": ["reviewable_but_not_deck_ready"] * 5,
                    "workbook_readiness_gate_status": ["WARN"] * 5,
                    "workbook_usage_restriction": ["Review material only. Not a final persona asset."] * 5,
                    "persona_size": [4490, 2570, 807, 598, 539],
                    "share_of_core_labeled": [47.6, 27.2, 8.5, 6.3, 5.7],
                    "share_of_all_labeled": [40.0, 23.0, 7.0, 5.0, 4.7],
                    "denominator_type": ["persona_core_labeled_rows"] * 5,
                    "denominator_value": [55] * 5,
                    "base_promotion_status": ["promoted_persona"] * 5,
                    "promoted_candidate_persona": [True] * 5,
                    "workbook_review_visible": [True, True, True, True, False],
                    "promotion_status": ["promoted_persona"] * 5,
                    "promotion_grounding_status": ["promoted_and_grounded"] * 5,
                    "final_usable_persona": [True, True, True, False, False],
                    "production_ready_persona": [True, True, True, False, False],
                    "review_ready_persona": [False, False, False, True, False],
                    "readiness_tier": [
                        "production_ready_persona",
                        "production_ready_persona",
                        "production_ready_persona",
                        "review_ready_persona",
                        "blocked_or_constrained_candidate",
                    ],
                    "review_visibility_status": [
                        "production_ready_visible",
                        "production_ready_visible",
                        "production_ready_visible",
                        "review_ready_visible",
                        "blocked_not_review_ready",
                    ],
                    "review_ready_reason": ["", "", "", "Locally grounded and structurally supported, but workbook-global concentration policy keeps this persona review-ready only.", ""],
                    "blocked_reason": ["", "", "", "blocked from production-ready use by workbook-global concentration policy", "weak-source dominated or insufficiently cross-source robust | thin evidence"],
                    "workbook_policy_constraint": ["", "", "", "top_3_cluster_share_of_core_labeled", "top_3_cluster_share_of_core_labeled"],
                    "dominant_signature": ["a", "b", "c", "d", "e"],
                    "dominant_bottleneck": ["manual_reporting", "manual_reporting", "dashboard_mistrust", "data_quality", "manual_reporting"],
                    "dominant_analysis_goal": ["report_speed", "report_speed", "decision_confidence", "validate_numbers", "report_speed"],
                }
            )
            persona_summary_df = pd.DataFrame(
                {
                    "persona_id": ["persona_01", "persona_02", "persona_03", "persona_04", "persona_05"],
                    "workbook_readiness_state": ["reviewable_but_not_deck_ready"] * 5,
                    "workbook_readiness_gate_status": ["WARN"] * 5,
                    "workbook_usage_restriction": ["Review material only. Not a final persona asset."] * 5,
                    "persona_name": ["P1", "P2", "P3", "P4", "P5"],
                    "persona_profile_name": ["P1", "P2", "P3", "P4", "P5"],
                    "legacy_persona_name": ["", "", "", "", ""],
                    "persona_size": [4490, 2570, 807, 598, 539],
                    "share_of_core_labeled": [47.6, 27.2, 8.5, 6.3, 5.7],
                    "share_of_all_labeled": [40.0, 23.0, 7.0, 5.0, 4.7],
                    "denominator_type": ["persona_core_labeled_rows"] * 5,
                    "denominator_value": [55] * 5,
                    "min_cluster_size": [200] * 5,
                    "base_promotion_status": ["promoted_persona"] * 5,
                    "promoted_candidate_persona": [True] * 5,
                    "workbook_review_visible": [True, True, True, True, False],
                    "visibility_state": ["workbook_visible"] * 5,
                    "final_usable_persona": [True, True, True, False, False],
                    "production_ready_persona": [True, True, True, False, False],
                    "review_ready_persona": [False, False, False, True, False],
                    "readiness_tier": [
                        "production_ready_persona",
                        "production_ready_persona",
                        "production_ready_persona",
                        "review_ready_persona",
                        "blocked_or_constrained_candidate",
                    ],
                    "review_ready_reason": ["", "", "", "Locally grounded and structurally supported, but workbook-global concentration policy keeps this persona review-ready only.", ""],
                    "blocked_reason": ["", "", "", "blocked from production-ready use by workbook-global concentration policy", "weak-source dominated or insufficiently cross-source robust | thin evidence"],
                    "workbook_policy_constraint": ["", "", "", "top_3_cluster_share_of_core_labeled", "top_3_cluster_share_of_core_labeled"],
                    "review_visibility_status": [
                        "production_ready_visible",
                        "production_ready_visible",
                        "production_ready_visible",
                        "review_ready_visible",
                        "blocked_not_review_ready",
                    ],
                    "usability_state": ["final_usable", "final_usable", "final_usable", "not_final_usable", "not_final_usable"],
                    "deck_ready_persona": [False] * 5,
                    "deck_readiness_state": ["not_deck_ready"] * 5,
                    "reporting_readiness_status": ["review_only"] * 5,
                    "promotion_action": ["keep"] * 5,
                    "promotion_status": ["promoted_persona"] * 5,
                    "grounding_status": ["grounded"] * 5,
                    "promotion_grounding_status": ["promoted_and_grounded"] * 5,
                    "promotion_reason": ["meets floor"] * 5,
                    "grounding_reason": ["grounded example exists"] * 5,
                    "grounded_candidate_count": [2, 2, 2, 2, 1],
                    "weak_candidate_count": [0, 0, 0, 0, 1],
                    "selected_example_count": [2, 2, 1, 1, 1],
                    "fallback_selected_count": [0, 0, 0, 0, 0],
                    "one_line_summary": ["s1", "s2", "s3", "s4", "s5"],
                    "dominant_bottleneck": ["manual_reporting", "manual_reporting", "dashboard_mistrust", "data_quality", "manual_reporting"],
                    "main_workflow_context": ["reporting", "reporting", "validation", "validation", "reporting"],
                    "analysis_behavior": ["report_speed", "report_speed", "decision_confidence", "validate_numbers", "report_speed"],
                    "trust_explanation_need": ["high"] * 5,
                    "current_tool_dependency": ["spreadsheet_heavy"] * 5,
                    "primary_output_expectation": ["xlsx"] * 5,
                    "top_pain_points": ["pain"] * 5,
                    "representative_examples": ["example"] * 5,
                    "why_this_persona_matters": ["matters"] * 5,
                }
            )
            frames = assemble_workbook_frames(
                overview_df=overview_df,
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [100]}),
                source_distribution_df=pd.DataFrame(),
                taxonomy_summary_df=pd.DataFrame(),
                cluster_stats_df=cluster_stats_df,
                persona_summary_df=persona_summary_df,
                persona_axes_df=pd.DataFrame(),
                persona_needs_df=pd.DataFrame(),
                persona_cooccurrence_df=pd.DataFrame(),
                persona_examples_df=pd.DataFrame(
                    {
                        "persona_id": ["persona_04", "persona_05"],
                        "example_rank": [1, 1],
                        "example_text": ["reconcile report totals before signoff", "manual reporting burden"],
                    }
                ),
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["WARN"], "threshold": [""], "status": ["warn"], "level": ["warning"], "denominator_type": [""], "denominator_value": [""], "notes": [""]}),
                source_diagnostics_df=pd.DataFrame(),
                quality_failures_df=pd.DataFrame(),
                metric_glossary_df=pd.DataFrame(),
            )
            with warnings.catch_warnings(record=True) as captured:
                warnings.simplefilter("always")
                output = export_workbook_from_frames(root, frames)

            review_ready_warning_text = "\n".join(str(item.message) for item in captured)
            self.assertNotIn("review_ready_persona", review_ready_warning_text)
            self.assertNotIn("production_ready_persona", review_ready_warning_text)
            self.assertNotIn("readiness_tier", review_ready_warning_text)

            workbook = load_workbook(output, read_only=True)
            try:
                overview_rows = self._sheet_rows_by_first_column(workbook["overview"])
                persona_summary_rows = list(workbook["persona_summary"].iter_rows(values_only=True))
                cluster_stats_rows = list(workbook["cluster_stats"].iter_rows(values_only=True))
                readme_review_ready_copy = self._readme_value_by_label(workbook, "Review-ready personas")
                readme_threshold_copy = self._readme_value_by_label(workbook, "Threshold discipline")
            finally:
                workbook.close()

            self.assertEqual(overview_rows["final_usable_persona_count"][2], 3)
            self.assertEqual(overview_rows["production_ready_persona_count"][2], 3)
            self.assertEqual(overview_rows["review_ready_persona_count"][2], 1)
            self.assertEqual(overview_rows["blocked_or_constrained_persona_count"][2], 1)

            persona_headers = list(persona_summary_rows[0])
            cluster_headers = list(cluster_stats_rows[0])
            self.assertEqual(
                persona_headers[:12],
                [
                    "persona_id",
                    "persona_name",
                    "persona_profile_name",
                    "legacy_persona_name",
                    "persona_size",
                    "readiness_tier",
                    "production_ready_persona",
                    "review_ready_persona",
                    "review_visibility_status",
                    "review_ready_reason",
                    "blocked_reason",
                    "workbook_policy_constraint",
                ],
            )
            self.assertEqual(
                cluster_headers[:9],
                [
                    "persona_id",
                    "persona_size",
                    "readiness_tier",
                    "production_ready_persona",
                    "review_ready_persona",
                    "review_visibility_status",
                    "review_ready_reason",
                    "blocked_reason",
                    "workbook_policy_constraint",
                ],
            )

            persona_id_index = persona_headers.index("persona_id")
            readiness_index = persona_headers.index("readiness_tier")
            review_ready_index = persona_headers.index("review_ready_persona")
            production_ready_index = persona_headers.index("production_ready_persona")
            visible_index = persona_headers.index("review_visibility_status")
            ordered_personas = [row[persona_id_index] for row in persona_summary_rows[1:6]]
            self.assertEqual(ordered_personas, ["persona_01", "persona_02", "persona_03", "persona_04", "persona_05"])

            persona_row_map = {row[persona_id_index]: row for row in persona_summary_rows[1:] if row and row[persona_id_index]}
            self.assertEqual(persona_row_map["persona_04"][readiness_index], "review_ready_persona")
            self.assertEqual(persona_row_map["persona_04"][review_ready_index], True)
            self.assertEqual(persona_row_map["persona_04"][production_ready_index], False)
            self.assertEqual(persona_row_map["persona_04"][visible_index], "review_ready_visible")
            self.assertEqual(persona_row_map["persona_05"][readiness_index], "blocked_or_constrained_candidate")
            self.assertEqual(persona_row_map["persona_05"][review_ready_index], False)
            self.assertEqual(persona_row_map["persona_05"][visible_index], "blocked_not_review_ready")

            cluster_persona_id_index = cluster_headers.index("persona_id")
            cluster_readiness_index = cluster_headers.index("readiness_tier")
            cluster_review_index = cluster_headers.index("review_ready_persona")
            cluster_row_map = {row[cluster_persona_id_index]: row for row in cluster_stats_rows[1:] if row and row[cluster_persona_id_index]}
            self.assertEqual(cluster_row_map["persona_04"][cluster_readiness_index], "review_ready_persona")
            self.assertEqual(cluster_row_map["persona_04"][cluster_review_index], True)
            self.assertEqual(cluster_row_map["persona_05"][cluster_readiness_index], "blocked_or_constrained_candidate")

            self.assertIn("not included in final usable persona count", str(readme_review_ready_copy))
            self.assertIn("does not relax workbook policy or production-ready thresholds", str(readme_threshold_copy))

    def test_export_uses_explicit_row_based_headers_for_counts_and_source_distribution(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame({"metric": ["raw_record_rows"], "value": [12]}),
                counts_df=pd.DataFrame(
                    {
                        "metric": ["raw_record_rows", "valid_candidate_rows", "labeled_episode_rows"],
                        "count": [12, 8, 5],
                        "denominator_type": ["raw_record_rows", "valid_candidate_rows", "labeled_episode_rows"],
                        "denominator_value": [12, 8, 5],
                        "definition": ["raw rows", "valid rows", "labeled rows"],
                    }
                ),
                source_distribution_df=pd.DataFrame(
                    {
                        "source": ["reddit"],
                        "raw_count": [12],
                        "normalized_count": [10],
                        "valid_count": [8],
                        "prefiltered_valid_count": [4],
                        "episode_count": [5],
                        "labeled_count": [5],
                        "share_of_labeled": [100.0],
                        "denominator_type": ["labeled_episode_rows"],
                        "denominator_value": [5],
                    }
                ),
                taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"], "why_it_matters": ["x"], "allowed_values_or_logic": ["a"], "evidence_fields": ["b"]}),
                cluster_stats_df=pd.DataFrame(),
                persona_summary_df=pd.DataFrame(),
                persona_axes_df=pd.DataFrame(),
                persona_needs_df=pd.DataFrame(),
                persona_cooccurrence_df=pd.DataFrame(),
                persona_examples_df=pd.DataFrame(),
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"], "threshold": [""], "status": ["pass"], "level": ["pass"], "denominator_type": [""], "denominator_value": [""], "notes": [""]}),
                source_diagnostics_df=pd.DataFrame(),
                quality_failures_df=pd.DataFrame(),
                metric_glossary_df=pd.DataFrame(),
            )
            output = export_workbook_from_frames(root, frames)
            workbook = load_workbook(output, read_only=True)
            try:
                counts_headers = [cell.value for cell in next(workbook["counts"].iter_rows(min_row=1, max_row=1))]
                source_headers = [cell.value for cell in next(workbook["source_distribution"].iter_rows(min_row=1, max_row=1))]
            finally:
                workbook.close()

            self.assertIn("metric_key", counts_headers)
            self.assertIn("row_count", counts_headers)
            self.assertIn("raw_record_rows_for_source", source_headers)
            self.assertIn("valid_candidate_rows_for_source", source_headers)
            self.assertIn("labeled_episode_rows_for_source", source_headers)
            self.assertIn("share_of_labeled_episode_rows_pct", source_headers)


if __name__ == "__main__":
    unittest.main()
