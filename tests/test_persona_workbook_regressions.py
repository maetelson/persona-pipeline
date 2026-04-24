"""Regression tests for persona workbook generator contracts."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.analysis.diagnostics import build_quality_failures, build_source_diagnostics, finalize_quality_checks
from src.analysis.persona_service import _apply_persona_name_policy, _build_cluster_stats_df
from src.analysis.quality_status import build_quality_metrics, evaluate_quality_status, flatten_quality_status_result
from src.analysis.stage_service import _annotate_persona_readiness_frame, _apply_workbook_promotion_constraints, _build_final_overview_df
from src.analysis.summary import build_quality_checks_df
from src.analysis.workbook_bundle import assemble_workbook_frames, validate_workbook_frames
from src.exporters.xlsx_exporter import export_workbook_from_frames
from src.utils.pipeline_schema import (
    DENOMINATOR_PERSONA_CORE_LABELED_ROWS,
    DENOMINATOR_RAW_RECORD_ROWS,
    PIPELINE_STAGE_METRIC_NAMES,
    WORKBOOK_SHEET_NAMES,
)


class PersonaWorkbookRegressionTests(unittest.TestCase):
    """Verify workbook generator policy contracts do not regress."""

    def test_share_columns_match_stated_denominator(self) -> None:
        frames = assemble_workbook_frames(
            overview_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"]}),
            counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [1]}),
            source_distribution_df=pd.DataFrame({"source": ["reddit"], "share_of_labeled": [100.0]}),
            taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"]}),
            cluster_stats_df=pd.DataFrame(
                {
                    "persona_id": ["persona_01"],
                    "persona_size": [3],
                    "share_of_core_labeled": [75.0],
                    "share_of_all_labeled": [30.0],
                    "denominator_type": [DENOMINATOR_PERSONA_CORE_LABELED_ROWS],
                    "denominator_value": [4],
                    "base_promotion_status": ["promoted_persona"],
                    "promotion_status": ["promoted_persona"],
                    "grounding_status": ["grounded"],
                    "promotion_grounding_status": ["promoted_and_grounded"],
                    "promotion_reason": ["meets floor"],
                    "grounding_reason": ["grounded example exists"],
                    "grounded_candidate_count": [2],
                    "weak_candidate_count": [0],
                    "selected_example_count": [1],
                    "fallback_selected_count": [0],
                    "dominant_signature": ["workflow_stage=reporting"],
                    "dominant_bottleneck": ["manual_reporting"],
                    "dominant_analysis_goal": ["report_speed"],
                }
            ),
            persona_summary_df=pd.DataFrame(
                {
                    "persona_id": ["persona_01"],
                    "persona_name": ["Reporting Operator"],
                    "persona_size": [3],
                    "share_of_core_labeled": [75.0],
                    "share_of_all_labeled": [30.0],
                    "denominator_type": [DENOMINATOR_PERSONA_CORE_LABELED_ROWS],
                    "denominator_value": [4],
                    "min_cluster_size": [2],
                    "base_promotion_status": ["promoted_persona"],
                    "promotion_status": ["promoted_persona"],
                    "grounding_status": ["grounded"],
                    "promotion_grounding_status": ["promoted_and_grounded"],
                    "promotion_reason": ["meets floor"],
                    "grounding_reason": ["grounded example exists"],
                    "grounded_candidate_count": [2],
                    "weak_candidate_count": [0],
                    "selected_example_count": [1],
                    "fallback_selected_count": [0],
                    "one_line_summary": ["summary"],
                    "dominant_bottleneck": ["manual_reporting"],
                    "main_workflow_context": ["reporting"],
                    "analysis_behavior": ["report_speed"],
                    "trust_explanation_need": ["high"],
                    "current_tool_dependency": ["spreadsheet_heavy"],
                    "primary_output_expectation": ["xlsx"],
                    "top_pain_points": ["rework"],
                    "representative_examples": ["example"],
                    "why_this_persona_matters": ["matters"],
                }
            ),
            persona_axes_df=pd.DataFrame(),
            persona_needs_df=pd.DataFrame(),
            persona_cooccurrence_df=pd.DataFrame(),
            persona_examples_df=pd.DataFrame(),
            quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"], "threshold": [""], "status": ["pass"], "level": ["pass"], "denominator_type": [""], "denominator_value": [""], "notes": [""]}),
            source_diagnostics_df=pd.DataFrame(),
            quality_failures_df=pd.DataFrame(),
            metric_glossary_df=pd.DataFrame(),
        )

        messages = validate_workbook_frames(frames)
        self.assertFalse(any(message.startswith("share denominator mismatch:") for message in messages))
        self.assertFalse(any(message.startswith("forbidden generic share column:") for message in messages))

    def test_largest_cluster_share_uses_persona_core_denominator(self) -> None:
        persona_source_df = pd.DataFrame(
            {
                "episode_id": ["e1", "e2", "e3", "e4"],
                "persona_id": ["persona_01", "persona_01", "persona_01", "persona_02"],
                "bottleneck_type": ["manual_reporting", "manual_reporting", "manual_reporting", "data_quality"],
                "analysis_goal": ["report_speed", "report_speed", "report_speed", "validate_numbers"],
            }
        )
        cluster_policy = {
            "min_cluster_size": 2,
            "status_by_persona": {
                "persona_01": {"status": "promoted_persona", "reason": "meets floor"},
                "persona_02": {"status": "exploratory_bucket", "reason": "below floor"},
            },
        }

        cluster_stats_df = _build_cluster_stats_df(
            persona_source_df=persona_source_df,
            axis_names=["bottleneck_type", "analysis_goal"],
            total_labeled_records=10,
            persona_core_labeled_records=4,
            cluster_policy=cluster_policy,
        )
        labeled_df = pd.DataFrame(
            {
                "episode_id": [f"l{i}" for i in range(10)],
                "persona_core_eligible": [True, True, True, True, False, False, False, False, False, False],
                "role_codes": ["analyst"] * 10,
                "question_codes": ["reporting"] * 10,
                "pain_codes": ["manual_reporting"] * 10,
                "output_codes": ["O_XLSX"] * 10,
            }
        )

        metrics = build_quality_metrics(
            stage_counts={
                "raw_record_rows": 10,
                "normalized_post_rows": 10,
                "valid_candidate_rows": 10,
                "prefiltered_valid_rows": 10,
                "episode_rows": 10,
                "labeled_episode_rows": 10,
            },
            labeled_df=labeled_df,
            source_stage_counts_df=pd.DataFrame({"source": ["reddit"], "raw_record_count": [10], "labeled_episode_count": [10]}),
            cluster_stats_df=cluster_stats_df,
            persona_examples_df=pd.DataFrame(),
            cluster_profiles=[],
        )

        self.assertEqual(float(metrics["largest_cluster_share_of_core_labeled"]), 75.0)

    def test_source_diagnostics_keep_same_grain_funnels_bounded_and_bridge_metrics_explicit(self) -> None:
        stage_counts_df = pd.DataFrame(
            {
                "source": ["reddit"],
                "raw_record_count": [5],
                "normalized_post_count": [5],
                "valid_post_count": [4],
                "prefiltered_valid_post_count": [2],
                "episode_count": [3],
                "labeled_episode_count": [3],
                "labelable_episode_count": [2],
                "effective_diversity_contribution": [0.6],
                "promoted_persona_episode_count": [2],
                "dominant_invalid_reason": ["reason_unavailable"],
                "dominant_prefilter_reason": ["reason_unavailable"],
                "valid_retention_reason": ["healthy_valid_post_retention"],
                "valid_retention_level": ["pass"],
                "prefilter_retention_reason": ["healthy_prefilter_retention"],
                "prefilter_retention_level": ["pass"],
                "episode_yield_reason": ["healthy_episode_yield"],
                "episode_yield_level": ["pass"],
                "labelable_coverage_reason": ["healthy_labelable_coverage"],
                "labelable_coverage_level": ["pass"],
                "grounding_contribution_reason": ["healthy_grounding_contribution"],
                "grounding_contribution_level": ["pass"],
                "concentration_risk_reason": ["concentration_risk_clear"],
                "concentration_risk_level": ["pass"],
                "diversity_contribution_reason": ["weak_diversity_contribution"],
                "diversity_contribution_level": ["warning"],
                "failure_reason_top": ["weak_diversity_contribution"],
                "failure_level": ["warning"],
                "recommended_seed_set": [""],
            }
        )

        diagnostics_df = build_source_diagnostics(stage_counts_df)

        bounded_pct = diagnostics_df[diagnostics_df["metric_type"].astype(str).eq("percentage")]
        self.assertTrue(pd.to_numeric(bounded_pct["metric_value"], errors="coerce").between(0.0, 100.0).all())

        bridge_rows = diagnostics_df[
            diagnostics_df["grain"].astype(str).eq("mixed_grain_bridge")
            & diagnostics_df["row_kind"].astype(str).eq("metric")
        ]
        self.assertFalse(bridge_rows["metric_name"].astype(str).str.contains("rate|share|survival", case=False, regex=True).any())
        self.assertTrue(bridge_rows["bounded_range"].astype(str).eq("unbounded_ratio").all())
        self.assertEqual(float(bridge_rows.loc[bridge_rows["metric_name"] == "episodes_per_prefiltered_valid_post", "metric_value"].iloc[0]), 1.5)

    def test_overview_cannot_hide_critical_status_reasons(self) -> None:
        flattened = {
            "persona_readiness_state": "exploratory_only",
            "persona_readiness_label": "Hypothesis Material",
            "persona_asset_class": "hypothesis_material",
            "persona_readiness_gate_status": "FAIL",
            "persona_completion_claim_allowed": False,
            "persona_usage_restriction": "Hypothesis material only. Not a final persona asset.",
            "persona_readiness_summary": "Exploratory workbook only.",
            "persona_readiness_blockers": "overall_unknown_ratio<=0.35 | promoted_persona_example_coverage_pct>=80.0",
            "persona_readiness_rule": "exploratory_only below reviewable thresholds",
            "overall_status": "FAIL",
            "quality_flag": "UNSTABLE",
            "quality_flag_rule": "UNSTABLE if any axis status is FAIL; EXPLORATORY if no FAIL and any axis status is WARN; otherwise OK.",
            "composite_reason_keys": "overall_unknown_critical | promoted_persona_examples_missing",
            "core_clustering_status": "WARN",
            "source_diversity_status": "OK",
            "example_grounding_status": "FAIL",
            "overall_unknown_status": "FAIL",
            "core_unknown_status": "OK",
            "core_coverage_status": "WARN",
            "effective_source_diversity_status": "OK",
            "source_concentration_status": "OK",
            "largest_cluster_dominance_status": "WARN",
            "grounding_coverage_status": "FAIL",
            "persona_core_coverage_of_all_labeled_pct": 61.2,
            "persona_core_unknown_ratio": 0.069204,
            "overall_unknown_ratio": 0.430085,
            "effective_labeled_source_count": 9.6,
            "largest_cluster_share_of_core_labeled": 55.0,
            "largest_labeled_source_share_pct": 48.7,
            "promoted_candidate_persona_count": 3,
            "promotion_visibility_persona_count": 3,
            "headline_persona_count": 2,
            "final_usable_persona_count": 2,
            "deck_ready_persona_count": 2,
            "promoted_persona_example_coverage_pct": 66.7,
            "promoted_persona_grounded_count": 2,
            "promoted_persona_weakly_grounded_count": 0,
            "promoted_persona_ungrounded_count": 1,
            "promoted_personas_weakly_grounded": "",
            "promoted_personas_missing_examples": "persona_02",
            "min_cluster_size": 24,
        }
        overview_df = _build_final_overview_df(
            axis_names=[{"axis_name": "workflow_stage"}, {"axis_name": "analysis_goal"}],
            quality_checks=flattened,
            stage_counts={
                "raw_record_rows": 7739,
                "normalized_post_rows": 7739,
                "valid_candidate_rows": 3438,
                "prefiltered_valid_rows": 783,
                "episode_rows": 472,
                "labeled_episode_rows": 472,
            },
            persona_core_labeled_rows=289,
            cluster_stats_df=pd.DataFrame(
                {
                    "base_promotion_status": ["promoted_candidate_persona", "promoted_candidate_persona", "promoted_candidate_persona", "exploratory_bucket"],
                    "promotion_status": ["promoted_persona", "promoted_persona", "review_visible_persona", "exploratory_bucket"],
                    "workbook_review_visible": [True, True, True, False],
                    "promotion_grounding_status": ["promoted_and_grounded", "promoted_and_grounded", "promoted_but_ungrounded", "exploratory_bucket"],
                    "final_usable_persona": [True, True, False, False],
                }
            ),
        )
        lookup = dict(zip(overview_df["metric"], overview_df["value"]))

        self.assertEqual(str(lookup["quality_flag"]), "UNSTABLE")
        self.assertEqual(str(lookup["persona_readiness_state"]), "exploratory_only")
        self.assertEqual(str(lookup["persona_asset_class"]), "hypothesis_material")
        self.assertEqual(str(lookup["persona_completion_claim_allowed"]), "False")
        self.assertEqual(str(lookup["overall_status"]), "FAIL")
        self.assertIn("promoted_persona_examples_missing", str(lookup["composite_reason_keys"]))
        self.assertEqual(str(lookup["example_grounding_status"]), "FAIL")
        self.assertNotIn("persona_count", lookup)
        self.assertEqual(float(lookup["promotion_visibility_persona_count"]), 3.0)
        self.assertEqual(float(lookup["headline_persona_count"]), 2.0)
        self.assertEqual(float(lookup["final_usable_persona_count"]), 2.0)
        self.assertEqual(float(lookup["deck_ready_persona_count"]), 2.0)
        self.assertEqual(float(lookup["promoted_persona_ungrounded_count"]), 1.0)

    def test_persona_readiness_gate_marks_unstable_workbook_as_hypothesis_material(self) -> None:
        evaluated = evaluate_quality_status(
            {
                "overall_unknown_ratio": 0.3881,
                "persona_core_coverage_of_all_labeled_pct": 68.4,
                "largest_source_influence_share_pct": 49.0,
                "fragile_tail_share_of_core_labeled": 0.06,
                "promoted_persona_example_coverage_pct": 50.0,
                "final_usable_persona_count": 2,
            }
        )
        flattened = flatten_quality_status_result(evaluated)

        self.assertEqual(flattened["persona_readiness_state"], "exploratory_only")
        self.assertEqual(flattened["persona_readiness_label"], "Hypothesis Material")
        self.assertEqual(flattened["persona_asset_class"], "hypothesis_material")
        self.assertEqual(flattened["persona_readiness_gate_status"], "FAIL")
        self.assertFalse(flattened["persona_completion_claim_allowed"])
        self.assertIn("persona_core_coverage_of_all_labeled_pct>=70.0", flattened["persona_readiness_blockers"])
        self.assertIn("promoted_persona_example_coverage_pct>=80.0", flattened["persona_readiness_blockers"])
        self.assertIn("largest_source_influence_share_pct<=45.0", flattened["persona_readiness_blockers"])

    def test_persona_readiness_gate_allows_final_asset_only_at_deck_ready_or_above(self) -> None:
        evaluated = evaluate_quality_status(
            {
                "overall_unknown_ratio": 0.2,
                "persona_core_coverage_of_all_labeled_pct": 80.0,
                "effective_balanced_source_count": 6.0,
                "largest_labeled_source_share_pct": 34.0,
                "largest_source_influence_share_pct": 34.9,
                "weak_source_cost_center_count": 0,
                "largest_cluster_share_of_core_labeled": 34.0,
                "top_3_cluster_share_of_core_labeled": 70.0,
                "micro_cluster_count": 0,
                "thin_evidence_cluster_count": 0,
                "min_cluster_separation": 0.13,
                "fragile_tail_share_of_core_labeled": 0.08,
                "promoted_persona_example_coverage_pct": 100.0,
                "final_usable_persona_count": 3,
            }
        )
        flattened = flatten_quality_status_result(evaluated)

        self.assertEqual(flattened["persona_readiness_state"], "deck_ready")
        self.assertEqual(flattened["persona_asset_class"], "final_persona_asset")
        self.assertEqual(flattened["persona_readiness_gate_status"], "OK")
        self.assertTrue(flattened["persona_completion_claim_allowed"])

    def test_promotion_constraint_keeps_strong_third_persona_when_not_weak_source_linked(self) -> None:
        cluster_stats_df = pd.DataFrame(
            {
                "persona_id": ["persona_01", "persona_02", "persona_03"],
                "promotion_status": ["promoted_persona", "promoted_persona", "promoted_persona"],
                "share_of_core_labeled": [48.1, 27.1, 8.6],
                "promotion_score": [0.79, 0.78, 0.70],
                "cross_source_robustness_score": [0.84, 0.80, 0.81],
                "selected_example_count": [6, 2, 1],
                "bundle_episode_count": [120, 95, 20],
                "persona_size": [4498, 2537, 803],
                "structural_support_status": ["structurally_supported", "structurally_supported", "structurally_supported"],
                "grounding_status": ["grounded_single", "grounded_single", "grounded_single"],
            }
        )
        persona_summary_df = cluster_stats_df.copy()
        persona_summary_df["promotion_grounding_status"] = "promoted_and_grounded"
        persona_summary_df["final_usable_persona"] = True
        persona_summary_df["deck_ready_persona"] = True
        persona_summary_df["workbook_review_visible"] = True
        persona_summary_df["promotion_action"] = "remain_promoted"
        audit_df = persona_summary_df.copy()
        outputs, summary = _apply_workbook_promotion_constraints(
            persona_service_outputs={
                "cluster_stats_df": cluster_stats_df,
                "persona_summary_df": persona_summary_df,
                "persona_promotion_grounding_audit_df": audit_df,
                "persona_assignments_df": pd.DataFrame(
                    {
                        "episode_id": ["e1", "e2", "e3", "e4", "e5", "e6"],
                        "persona_id": ["persona_01", "persona_01", "persona_02", "persona_02", "persona_03", "persona_03"],
                    }
                ),
            },
            clustering_episodes_df=pd.DataFrame(
                {
                    "episode_id": ["e1", "e2", "e3", "e4", "e5", "e6"],
                    "source": ["reddit", "reddit", "reddit", "reddit", "power_bi_community", "metabase_discussions"],
                }
            ),
            source_balance_audit_df=pd.DataFrame(
                {
                    "source": ["google_developer_forums", "klaviyo_community", "power_bi_community", "metabase_discussions"],
                    "weak_source_cost_center": [True, True, False, False],
                    "blended_influence_share_pct": [30.6, 12.0, 18.0, 16.0],
                }
            ),
        )
        result = outputs["cluster_stats_df"].set_index("persona_id")
        self.assertIn(summary["promotion_constraint_status"], {"constrained", "constrained_no_borderline_match"})
        self.assertEqual(str(result.loc["persona_03", "promotion_status"]), "promoted_persona")

    def test_persona_readiness_is_capped_below_deck_ready_when_quality_status_warns(self) -> None:
        evaluated = evaluate_quality_status(
            {
                "overall_unknown_ratio": 0.2,
                "persona_core_coverage_of_all_labeled_pct": 80.0,
                "effective_balanced_source_count": 5.5,
                "largest_labeled_source_share_pct": 34.0,
                "largest_source_influence_share_pct": 34.9,
                "weak_source_cost_center_count": 0,
                "largest_cluster_share_of_core_labeled": 34.0,
                "fragile_tail_share_of_core_labeled": 0.08,
                "promoted_persona_example_coverage_pct": 100.0,
                "final_usable_persona_count": 3,
                "top_3_cluster_share_of_core_labeled": 0.834,
                "micro_cluster_count": 0,
                "thin_evidence_cluster_count": 0,
                "min_cluster_separation": 0.13,
            }
        )
        flattened = flatten_quality_status_result(evaluated)

        self.assertEqual(flattened["overall_status"], "WARN")
        self.assertEqual(flattened["quality_flag"], "EXPLORATORY")
        self.assertEqual(flattened["persona_readiness_state"], "reviewable_but_not_deck_ready")
        self.assertEqual(flattened["persona_asset_class"], "reviewable_draft")
        self.assertFalse(flattened["persona_completion_claim_allowed"])
        self.assertIn("overall_status=WARN keeps workbook below deck_ready", flattened["persona_readiness_blockers"])

    def test_persona_summary_rows_carry_workbook_readiness_gate(self) -> None:
        quality_checks = {
            "persona_readiness_state": "exploratory_only",
            "persona_readiness_gate_status": "FAIL",
            "persona_usage_restriction": "Hypothesis material only. Not a final persona asset and not safe for review sign-off, deck-ready use, or production persona use.",
        }
        persona_summary_df = pd.DataFrame(
            {
                "persona_id": ["persona_01"],
                "persona_name": ["Analyst"],
                "persona_size": [4],
            }
        )

        annotated = _annotate_persona_readiness_frame(persona_summary_df, quality_checks)

        self.assertEqual(str(annotated.loc[0, "workbook_readiness_state"]), "exploratory_only")
        self.assertEqual(str(annotated.loc[0, "workbook_readiness_gate_status"]), "FAIL")
        self.assertIn("Not a final persona asset", str(annotated.loc[0, "workbook_usage_restriction"]))

    def test_promoted_personas_without_examples_are_explicitly_flagged(self) -> None:
        labeled_df = pd.DataFrame(
            {
                "episode_id": ["e1", "e2", "e3", "e4"],
                "persona_core_eligible": [True, True, True, True],
                "role_codes": ["analyst"] * 4,
                "question_codes": ["reporting"] * 4,
                "pain_codes": ["manual_reporting"] * 4,
                "output_codes": ["O_XLSX"] * 4,
            }
        )
        cluster_stats_df = pd.DataFrame(
            {
                "persona_id": ["persona_01", "persona_02"],
                "persona_size": [3, 1],
                "share_of_core_labeled": [75.0, 25.0],
                "share_of_all_labeled": [75.0, 25.0],
                "denominator_type": [DENOMINATOR_PERSONA_CORE_LABELED_ROWS, DENOMINATOR_PERSONA_CORE_LABELED_ROWS],
                "denominator_value": [4, 4],
                "base_promotion_status": ["promoted_candidate_persona", "promoted_candidate_persona"],
                "promotion_status": ["promoted_persona", "review_visible_persona"],
                "workbook_review_visible": [True, True],
                "promotion_grounding_status": ["promoted_and_grounded", "promoted_but_ungrounded"],
                "final_usable_persona": [True, False],
            }
        )
        persona_examples_df = pd.DataFrame(
            {
                "persona_id": ["persona_01"],
                "example_rank": [1],
                "grounded_text": ["Strong example"],
            }
        )

        metrics = build_quality_metrics(
            stage_counts={
                "raw_record_rows": 4,
                "normalized_post_rows": 4,
                "valid_candidate_rows": 4,
                "prefiltered_valid_rows": 4,
                "episode_rows": 4,
                "labeled_episode_rows": 4,
            },
            labeled_df=labeled_df,
            source_stage_counts_df=pd.DataFrame({"source": ["reddit"], "raw_record_count": [4], "labeled_episode_count": [4]}),
            cluster_stats_df=cluster_stats_df,
            persona_examples_df=persona_examples_df,
            cluster_profiles=[],
        )
        flattened = evaluate_quality_status(metrics)
        overview_df = _build_final_overview_df(
            axis_names=[{"axis_name": "workflow_stage"}],
            quality_checks=flattened["metrics"] | {
                "overall_status": flattened["composite_status"],
                "quality_flag": flattened["quality_flag"],
                "quality_flag_rule": flattened["quality_flag_rule"],
                "composite_reason_keys": " | ".join(flattened["composite_reason_keys"]),
                "core_clustering_status": flattened["groups"]["core_clustering"]["status"],
                "source_diversity_status": flattened["groups"]["source_diversity"]["status"],
                "example_grounding_status": flattened["groups"]["example_grounding"]["status"],
                **{f"{axis}_status": payload["status"] for axis, payload in flattened["axes"].items()},
            },
            stage_counts={
                "raw_record_rows": 4,
                "normalized_post_rows": 4,
                "valid_candidate_rows": 4,
                "prefiltered_valid_rows": 4,
                "episode_rows": 4,
                "labeled_episode_rows": 4,
            },
            persona_core_labeled_rows=4,
            cluster_stats_df=cluster_stats_df,
        )
        lookup = dict(zip(overview_df["metric"], overview_df["value"]))

        self.assertEqual(float(metrics["promoted_persona_example_coverage_pct"]), 50.0)
        self.assertEqual(int(metrics["promotion_visibility_persona_count"]), 2)
        self.assertEqual(int(metrics["headline_persona_count"]), 1)
        self.assertEqual(int(metrics["final_usable_persona_count"]), 1)
        self.assertEqual(int(metrics["deck_ready_persona_count"]), 1)
        self.assertEqual(int(metrics["promoted_persona_grounding_failure_count"]), 1)
        self.assertEqual(int(metrics["selected_example_grounding_issue_count"]), 0)
        self.assertEqual(str(metrics["promoted_personas_missing_examples"]), "persona_02")
        self.assertEqual(str(lookup["promoted_personas_missing_examples"]), "persona_02")

    def test_finalize_quality_checks_zeroes_deck_ready_persona_count_below_deck_ready(self) -> None:
        evaluated = evaluate_quality_status(
            {
                "overall_unknown_ratio": 0.2,
                "persona_core_coverage_of_all_labeled_pct": 80.0,
                "effective_balanced_source_count": 5.5,
                "largest_labeled_source_share_pct": 34.0,
                "largest_source_influence_share_pct": 34.9,
                "weak_source_cost_center_count": 0,
                "largest_cluster_share_of_core_labeled": 34.0,
                "fragile_tail_share_of_core_labeled": 0.08,
                "promoted_persona_example_coverage_pct": 100.0,
                "final_usable_persona_count": 3,
                "top_3_cluster_share_of_core_labeled": 0.834,
                "micro_cluster_count": 0,
                "thin_evidence_cluster_count": 0,
                "min_cluster_separation": 0.13,
                "deck_ready_persona_count": 3,
            }
        )
        finalized = finalize_quality_checks(evaluated)

        self.assertEqual(finalized["persona_readiness_state"], "reviewable_but_not_deck_ready")
        self.assertEqual(int(finalized["deck_ready_persona_count"]), 0)

    def test_persona_name_policy_adds_distinctive_suffixes_for_near_duplicates(self) -> None:
        summary_df = pd.DataFrame(
            {
                "persona_id": ["persona_01", "persona_02"],
                "promotion_status": ["promoted_persona", "promoted_persona"],
                "legacy_persona_name": ["Analyst Reporting Blocked by Spreadsheet Rework"] * 2,
                "persona_profile_name": ["Analyst Reporting Operator"] * 2,
                "recurring_job_to_be_done": ["deliver_recurring_reporting_without_manual_repackaging"] * 2,
                "primary_bottleneck": ["manual_reporting", "data_quality"],
                "trust_failure_mode": ["not_explainable_to_stakeholders", "signoff_pressure_on_metric_quality"],
                "functional_context": ["reporting_and_performance_management", "metric_validation_and_signoff"],
                "user_role_family": ["analyst_operator", "analyst_operator"],
            }
        )

        named = _apply_persona_name_policy(summary_df)

        self.assertEqual(len(set(named["persona_name"].astype(str).tolist())), 2)
        self.assertTrue(all("Analyst Reporting Operator" not in value for value in named["persona_name"].astype(str).tolist()))

    def test_persona_name_policy_rewrites_exploratory_role_heavy_names_into_residual_signatures(self) -> None:
        summary_df = pd.DataFrame(
            {
                "persona_id": ["persona_03", "persona_04"],
                "promotion_status": ["exploratory_bucket", "exploratory_bucket"],
                "legacy_persona_name": [
                    "Analyst Workflow Blocked by Spreadsheet Rework",
                    "Analyst Workflow Blocked by Explanation Gaps",
                ],
                "persona_profile_name": ["Analyst Reporting Operator", "Analyst Analysis Operator"],
                "recurring_job_to_be_done": [
                    "deliver_recurring_reporting_without_manual_repackaging",
                    "move_analysis_work_to_a_shareable_output",
                ],
                "primary_bottleneck": ["manual_reporting", "handoff_dependency"],
                "trust_failure_mode": [
                    "numbers_do_not_reconcile_or_feel_safe_to_share",
                    "context_is_not_explainable_without_manual_follow_up",
                ],
                "functional_context": [
                    "reporting_and_performance_management",
                    "analytics_workflow_execution",
                ],
                "expected_output_artifact": [
                    "stakeholder_ready_export_or_packaged_report",
                    "updated_dashboard_or_explanation_for_existing_dashboard",
                ],
                "typical_trigger_event": [
                    "scheduled_reporting_cycle_or_stakeholder_request",
                    "cross_team_follow_up_blocks_delivery",
                ],
                "workaround_pattern": [
                    "export_then_patch_in_spreadsheet",
                    "escalate_to_other_teams_for_context_and_validation",
                ],
            }
        )

        named = _apply_persona_name_policy(summary_df)
        values = named["persona_name"].astype(str).tolist()

        self.assertEqual(len(set(values)), 2)
        self.assertTrue(all("Analyst " not in value for value in values))
        self.assertTrue(any("Spreadsheet Patchwork" in value for value in values))
        self.assertTrue(any("Cross-Team Explanation" in value for value in values))

    def test_grounded_but_structurally_weak_persona_does_not_count_as_final_usable(self) -> None:
        labeled_df = pd.DataFrame(
            {
                "episode_id": ["e1", "e2", "e3", "e4"],
                "persona_core_eligible": [True, True, True, True],
                "role_codes": ["analyst"] * 4,
                "question_codes": ["reporting"] * 4,
                "pain_codes": ["manual_reporting"] * 4,
                "output_codes": ["O_XLSX"] * 4,
            }
        )
        cluster_stats_df = pd.DataFrame(
            {
                "persona_id": ["persona_01"],
                "persona_size": [4],
                "share_of_core_labeled": [100.0],
                "share_of_all_labeled": [100.0],
                "denominator_type": [DENOMINATOR_PERSONA_CORE_LABELED_ROWS],
                "denominator_value": [4],
                "base_promotion_status": ["promoted_candidate_persona"],
                "promotion_status": ["review_visible_persona"],
                "workbook_review_visible": [True],
                "structural_support_status": ["review_visible_only"],
                "grounding_status": ["grounded"],
                "promotion_grounding_status": ["grounded_but_structurally_weak"],
                "final_usable_persona": [False],
                "deck_ready_persona": [False],
            }
        )
        persona_examples_df = pd.DataFrame(
            {
                "persona_id": ["persona_01"],
                "example_rank": [1],
                "grounded_text": ["Strong example"],
            }
        )

        metrics = build_quality_metrics(
            stage_counts={
                "raw_record_rows": 4,
                "normalized_post_rows": 4,
                "valid_candidate_rows": 4,
                "prefiltered_valid_rows": 4,
                "episode_rows": 4,
                "labeled_episode_rows": 4,
            },
            labeled_df=labeled_df,
            source_stage_counts_df=pd.DataFrame({"source": ["reddit"], "raw_record_count": [4], "labeled_episode_count": [4]}),
            cluster_stats_df=cluster_stats_df,
            persona_examples_df=persona_examples_df,
            cluster_profiles=[],
        )

        self.assertEqual(int(metrics["promoted_persona_grounded_count"]), 1)
        self.assertEqual(int(metrics["promotion_visibility_persona_count"]), 1)
        self.assertEqual(int(metrics["final_usable_persona_count"]), 0)
        self.assertEqual(int(metrics["headline_persona_count"]), 0)

    def test_generation_keeps_required_sheets_and_workbook_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "config").mkdir(parents=True, exist_ok=True)
            (root / "config" / "export_schema.yaml").write_text("workbook_name: test.xlsx\n", encoding="utf-8")
            frames = assemble_workbook_frames(
                overview_df=pd.DataFrame(
                    {
                        "metric": [
                            "quality_flag",
                            "promoted_candidate_persona_count",
                            "promotion_visibility_persona_count",
                            "headline_persona_count",
                            "final_usable_persona_count",
                            "deck_ready_persona_count",
                        ],
                            "value": ["OK", 1, 1, 1, 1, 1],
                    }
                ),
                counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [1], "denominator_type": [DENOMINATOR_RAW_RECORD_ROWS], "denominator_value": [1], "definition": ["rows"]}),
                source_distribution_df=pd.DataFrame({"source": ["reddit"], "raw_count": [1], "normalized_count": [1], "valid_count": [1], "prefiltered_valid_count": [1], "episode_count": [1], "labeled_count": [1], "share_of_labeled": [100.0], "denominator_type": ["labeled_episode_rows"], "denominator_value": [1]}),
                taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"], "why_it_matters": ["x"], "allowed_values_or_logic": ["y"], "evidence_fields": ["z"]}),
                cluster_stats_df=pd.DataFrame({"persona_id": ["persona_01"], "persona_size": [1], "share_of_core_labeled": [100.0], "share_of_all_labeled": [100.0], "denominator_type": [DENOMINATOR_PERSONA_CORE_LABELED_ROWS], "denominator_value": [1], "min_cluster_size": [1], "base_promotion_status": ["promoted_candidate_persona"], "promoted_candidate_persona": [True], "workbook_review_visible": [True], "final_usable_persona": [True], "deck_ready_persona": [True], "reporting_readiness_status": ["final_usable_persona"], "promotion_status": ["promoted_persona"], "grounding_status": ["grounded"], "promotion_grounding_status": ["promoted_and_grounded"], "promotion_reason": ["meets floor"], "grounding_reason": ["grounded example exists"], "grounded_candidate_count": [1], "weak_candidate_count": [0], "selected_example_count": [1], "fallback_selected_count": [0], "dominant_signature": ["role=analyst"], "dominant_bottleneck": ["manual_reporting"], "dominant_analysis_goal": ["report_speed"]}),
                persona_summary_df=pd.DataFrame({"persona_id": ["persona_01"], "persona_name": ["Analyst"], "persona_size": [1], "share_of_core_labeled": [100.0], "share_of_all_labeled": [100.0], "denominator_type": [DENOMINATOR_PERSONA_CORE_LABELED_ROWS], "denominator_value": [1], "min_cluster_size": [1], "base_promotion_status": ["promoted_candidate_persona"], "promoted_candidate_persona": [True], "workbook_review_visible": [True], "final_usable_persona": [True], "deck_ready_persona": [True], "reporting_readiness_status": ["final_usable_persona"], "promotion_status": ["promoted_persona"], "grounding_status": ["grounded"], "promotion_grounding_status": ["promoted_and_grounded"], "promotion_reason": ["meets floor"], "grounding_reason": ["grounded example exists"], "grounded_candidate_count": [1], "weak_candidate_count": [0], "selected_example_count": [1], "fallback_selected_count": [0], "one_line_summary": ["summary"], "dominant_bottleneck": ["manual_reporting"], "main_workflow_context": ["reporting"], "analysis_behavior": ["report_speed"], "trust_explanation_need": ["high"], "current_tool_dependency": ["spreadsheet_heavy"], "primary_output_expectation": ["xlsx"], "top_pain_points": ["rework"], "representative_examples": ["example"], "why_this_persona_matters": ["matters"]}),
                persona_axes_df=pd.DataFrame({"persona_id": ["persona_01"], "axis_name": ["role"], "axis_value": ["analyst"], "count": [1], "pct_of_persona": [100.0]}),
                persona_needs_df=pd.DataFrame({"persona_id": ["persona_01"], "pain_or_need": ["rework"], "count": [1], "pct_of_persona": [100.0], "rank": [1]}),
                persona_cooccurrence_df=pd.DataFrame({"persona_id": ["persona_01"], "theme_a": ["a"], "theme_b": ["b"], "pair_count": [1], "pct_of_persona": [100.0], "rank": [1]}),
                persona_examples_df=pd.DataFrame({"persona_id": ["persona_01"], "example_rank": [1], "grounded_text": ["example"], "selection_strength": ["grounded"], "grounding_strength": ["grounded"], "fallback_selected": [False], "coverage_selection_reason": ["score_plus_diversity_policy"], "grounding_reason": ["grounded"], "why_selected": ["because"], "matched_axes": ["role=analyst"], "reason_selected": ["fit"], "quote_quality": ["usable"], "grounding_fit_score": [1.8], "mismatch_count": [0], "critical_mismatch_count": [0], "matched_axis_count": [4], "final_example_score": [9.2]}),
                quality_checks_df=pd.DataFrame({"metric": ["quality_flag"], "value": ["OK"], "threshold": [""], "status": ["pass"], "level": ["pass"], "denominator_type": [""], "denominator_value": [""], "notes": [""]}),
                source_diagnostics_df=pd.DataFrame({"source": ["reddit", "reddit"], "section": ["post_funnel", "diagnostic_reasons"], "row_kind": ["metric", "diagnostic"], "grain": ["post", "other"], "metric_name": ["valid_posts_per_normalized_post_pct", "top_failure_reason"], "metric_value": [100.0, "healthy_source_contribution"], "metric_type": ["percentage", "diagnostic_reason"], "denominator_metric": ["normalized_post_count", ""], "denominator_grain": ["post", ""], "denominator_value": [1, ""], "bounded_range": ["0-100_pct", ""], "is_same_grain_funnel": [True, False], "diagnostic_level": ["", "pass"], "metric_definition": ["definition", "definition"]}),
                quality_failures_df=pd.DataFrame({"metric": ["denominator_consistency_check"], "level": ["pass"], "value": ["explicit"], "threshold": [""], "passed": [True]}),
                metric_glossary_df=pd.DataFrame({"metric": ["quality_flag"], "denominator_type": [""], "definition": ["definition"]}),
            )

            output = export_workbook_from_frames(root, frames)
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(workbook.sheetnames[0], "readme")
                self.assertEqual(workbook.sheetnames[1:], WORKBOOK_SHEET_NAMES)
                overview_headers = [cell.value for cell in next(workbook["overview"].iter_rows(min_row=1, max_row=1))]
                source_headers = [cell.value for cell in next(workbook["source_diagnostics"].iter_rows(min_row=1, max_row=1))]
                persona_summary_headers = [cell.value for cell in next(workbook["persona_summary"].iter_rows(min_row=1, max_row=1))]
                persona_example_headers = [cell.value for cell in next(workbook["persona_examples"].iter_rows(min_row=1, max_row=1))]
            finally:
                workbook.close()

            self.assertIn("metric_key", overview_headers)
            self.assertIn("display_label", overview_headers)
            self.assertIn("metric_value", overview_headers)
            self.assertIn("metric_name", source_headers)
            self.assertIn("row_kind", source_headers)
            self.assertIn("row_grain", source_headers)
            self.assertIn("promotion_grounding_status", persona_summary_headers)
            self.assertIn("grounding_status", persona_summary_headers)
            self.assertIn("final_usable_persona", persona_summary_headers)
            self.assertIn("reporting_readiness_status", persona_summary_headers)
            self.assertIn("selected_example_strength", persona_example_headers)
            self.assertIn("example_grounding_strength", persona_example_headers)
            self.assertIn("fallback_selected", persona_example_headers)

    def test_validate_workbook_frames_rejects_ambiguous_persona_count_for_review_visible_personas(self) -> None:
        frames = assemble_workbook_frames(
            overview_df=pd.DataFrame({"metric": ["persona_count"], "value": [3]}),
            counts_df=pd.DataFrame({"metric": ["raw_record_rows"], "count": [7], "denominator_type": ["raw_record_rows"], "denominator_value": [7], "definition": ["raw"]}),
            source_distribution_df=pd.DataFrame({"source": ["reddit"], "share_of_labeled": [100.0]}),
            taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"]}),
            cluster_stats_df=pd.DataFrame(
                {
                    "persona_id": ["persona_01", "persona_02", "persona_03"],
                    "base_promotion_status": ["promoted_candidate_persona", "promoted_candidate_persona", "promoted_candidate_persona"],
                    "promotion_status": ["promoted_persona", "promoted_persona", "review_visible_persona"],
                    "workbook_review_visible": [True, True, True],
                    "promotion_grounding_status": ["promoted_and_grounded", "promoted_and_grounded", "promoted_but_ungrounded"],
                    "final_usable_persona": [True, True, False],
                }
            ),
            persona_summary_df=pd.DataFrame(),
            persona_axes_df=pd.DataFrame(),
            persona_needs_df=pd.DataFrame(),
            persona_cooccurrence_df=pd.DataFrame(),
            persona_examples_df=pd.DataFrame(),
            quality_checks_df=pd.DataFrame(),
            source_diagnostics_df=pd.DataFrame(),
            quality_failures_df=pd.DataFrame(),
            metric_glossary_df=pd.DataFrame({"metric": ["persona_count"], "denominator_type": [""], "definition": ["ambiguous"]}),
        )

        messages = validate_workbook_frames(frames)
        self.assertTrue(any(message == "ambiguous persona count metric: overview.persona_count" for message in messages))
        self.assertTrue(any(message == "ambiguous persona count metric: metric_glossary.persona_count" for message in messages))
        self.assertTrue(any(message.startswith("missing persona promotion metric: overview.promotion_visibility_persona_count") for message in messages))
        self.assertTrue(any(message.startswith("missing persona promotion metric: overview.headline_persona_count") for message in messages))

    def test_validate_workbook_frames_rejects_legacy_stage_aliases_and_stage_value_drift(self) -> None:
        frames = assemble_workbook_frames(
            overview_df=pd.DataFrame({"metric": ["raw_record_rows", "labeled_episode_rows"], "value": [7, 5]}),
            counts_df=pd.DataFrame({"metric": ["raw_record_rows", "labeled_episode_rows"], "count": [7, 5], "denominator_type": ["raw_record_rows", "labeled_episode_rows"], "denominator_value": [7, 5], "definition": ["raw", "labeled"]}),
            source_distribution_df=pd.DataFrame({"source": ["reddit"], "share_of_labeled": [100.0]}),
            taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"]}),
            cluster_stats_df=pd.DataFrame(),
            persona_summary_df=pd.DataFrame(),
            persona_axes_df=pd.DataFrame(),
            persona_needs_df=pd.DataFrame(),
            persona_cooccurrence_df=pd.DataFrame(),
            persona_examples_df=pd.DataFrame(),
            quality_checks_df=pd.DataFrame({"metric": ["raw_records", "labeled_episode_rows"], "value": [7, 6], "threshold": ["", ""], "status": ["pass", "pass"], "level": ["pass", "pass"], "denominator_type": ["raw_record_rows", "labeled_episode_rows"], "denominator_value": [7, 6], "notes": ["", ""]}),
            source_diagnostics_df=pd.DataFrame(),
            quality_failures_df=pd.DataFrame(),
            metric_glossary_df=pd.DataFrame(),
        )

        messages = validate_workbook_frames(frames)
        self.assertTrue(any(message.startswith("legacy stage metric alias: quality_checks.raw_records->raw_record_rows") for message in messages))
        self.assertTrue(any(message.startswith("stage metric mismatch: labeled_episode_rows") for message in messages))

    def test_canonical_stage_metrics_can_match_across_counts_overview_and_quality_checks(self) -> None:
        stage_rows = {metric: index + 1 for index, metric in enumerate(PIPELINE_STAGE_METRIC_NAMES)}
        counts_df = pd.DataFrame(
            {
                "metric": list(stage_rows.keys()),
                "count": list(stage_rows.values()),
                "denominator_type": list(stage_rows.keys()),
                "denominator_value": list(stage_rows.values()),
                "definition": [metric for metric in stage_rows],
            }
        )
        overview_df = pd.DataFrame({"metric": list(stage_rows.keys()), "value": list(stage_rows.values())})
        quality_checks_df = pd.DataFrame(
            {
                "metric": list(stage_rows.keys()),
                "value": list(stage_rows.values()),
                "threshold": [""] * len(stage_rows),
                "status": ["pass"] * len(stage_rows),
                "level": ["pass"] * len(stage_rows),
                "denominator_type": list(stage_rows.keys()),
                "denominator_value": list(stage_rows.values()),
                "notes": [""] * len(stage_rows),
            }
        )
        frames = assemble_workbook_frames(
            overview_df=overview_df,
            counts_df=counts_df,
            source_distribution_df=pd.DataFrame({"source": ["reddit"], "share_of_labeled": [100.0]}),
            taxonomy_summary_df=pd.DataFrame({"axis_name": ["role"]}),
            cluster_stats_df=pd.DataFrame(),
            persona_summary_df=pd.DataFrame(),
            persona_axes_df=pd.DataFrame(),
            persona_needs_df=pd.DataFrame(),
            persona_cooccurrence_df=pd.DataFrame(),
            persona_examples_df=pd.DataFrame(),
            quality_checks_df=quality_checks_df,
            source_diagnostics_df=pd.DataFrame(),
            quality_failures_df=pd.DataFrame(),
            metric_glossary_df=pd.DataFrame(),
        )
        messages = validate_workbook_frames(frames)
        self.assertFalse(any(message.startswith("legacy stage metric alias:") for message in messages))
        self.assertFalse(any(message.startswith("stage metric mismatch:") for message in messages))

    def test_overview_quality_checks_and_failures_share_grounding_narrative(self) -> None:
        flattened = {
            "overall_status": "FAIL",
            "quality_flag": "UNSTABLE",
            "quality_flag_rule": "UNSTABLE if any axis status is FAIL; EXPLORATORY if no FAIL and any axis status is WARN; otherwise OK.",
            "composite_reason_keys": "promoted_persona_examples_coverage_critical | promoted_persona_examples_missing",
            "core_clustering_status": "OK",
            "source_diversity_status": "OK",
            "example_grounding_status": "FAIL",
            "overall_unknown_status": "OK",
            "core_unknown_status": "OK",
            "core_coverage_status": "OK",
            "effective_source_diversity_status": "OK",
            "source_concentration_status": "OK",
            "largest_cluster_dominance_status": "OK",
            "grounding_coverage_status": "FAIL",
            "grounding_coverage_reason_keys": "promoted_persona_examples_coverage_critical | promoted_persona_examples_missing",
            "persona_core_coverage_of_all_labeled_pct": 90.0,
            "persona_core_unknown_ratio": 0.01,
            "overall_unknown_ratio": 0.02,
            "effective_labeled_source_count": 6.0,
            "largest_cluster_share_of_core_labeled": 40.0,
            "largest_labeled_source_share_pct": 30.0,
            "promoted_candidate_persona_count": 3,
            "promotion_visibility_persona_count": 3,
            "headline_persona_count": 2,
            "final_usable_persona_count": 2,
            "deck_ready_persona_count": 2,
            "promoted_persona_example_coverage_pct": 66.7,
            "promoted_persona_grounded_count": 2,
            "promoted_persona_weakly_grounded_count": 0,
            "promoted_persona_ungrounded_count": 1,
            "promoted_persona_grounding_failure_count": 1,
            "selected_example_grounding_issue_count": 0,
            "promoted_personas_missing_examples": "persona_02",
            "promoted_personas_weakly_grounded": "",
            "min_cluster_size": 24,
            "denominator_consistency": "explicit",
        }
        cluster_stats_df = pd.DataFrame(
            {
                "persona_id": ["persona_01", "persona_02", "persona_03"],
                "promotion_status": ["promoted_persona", "review_visible_persona", "promoted_persona"],
                "workbook_review_visible": [True, True, True],
                "promotion_grounding_status": ["promoted_and_grounded", "promoted_but_ungrounded", "promoted_and_grounded"],
                "final_usable_persona": [True, False, True],
            }
        )

        overview_df = _build_final_overview_df(
            axis_names=[{"axis_name": "workflow_stage"}],
            quality_checks=flattened,
            stage_counts={
                "raw_record_rows": 100,
                "normalized_post_rows": 100,
                "valid_candidate_rows": 80,
                "prefiltered_valid_rows": 30,
                "episode_rows": 20,
                "labeled_episode_rows": 20,
            },
            persona_core_labeled_rows=18,
            cluster_stats_df=cluster_stats_df,
        )
        quality_df = build_quality_checks_df(flattened)
        failures_df = build_quality_failures(flattened, pd.DataFrame(), cluster_stats_df, pd.DataFrame())

        overview_lookup = dict(zip(overview_df["metric"], overview_df["value"]))
        quality_lookup = {
            str(row["metric"]): row
            for _, row in quality_df.iterrows()
        }
        failure_lookup = {
            str(row["metric"]): row
            for _, row in failures_df.iterrows()
        }

        self.assertEqual(str(overview_lookup["grounding_coverage_status"]), "FAIL")
        self.assertEqual(float(overview_lookup["promotion_visibility_persona_count"]), 3.0)
        self.assertEqual(float(overview_lookup["headline_persona_count"]), 2.0)
        self.assertEqual(float(overview_lookup["final_usable_persona_count"]), 2.0)
        self.assertEqual(float(overview_lookup["promoted_persona_ungrounded_count"]), 1.0)

        self.assertEqual(str(quality_lookup["promoted_persona_grounding_failure_count"]["status"]), "fail")
        self.assertEqual(str(quality_lookup["promoted_persona_grounding_failure_count"]["level"]), "hard_fail")
        self.assertEqual(str(quality_lookup["selected_example_grounding_issue_count"]["status"]), "pass")
        self.assertIn("no selected representative examples", str(quality_lookup["selected_example_grounding_issue_count"]["notes"]))
        self.assertEqual(str(quality_lookup["promoted_persona_example_coverage_pct"]["status"]), "fail")

        self.assertEqual(str(failure_lookup["promoted_persona_grounding_gate"]["level"]), "hard_fail")
        self.assertEqual(int(failure_lookup["promoted_persona_grounding_gate"]["value"]), 1)
        self.assertEqual(str(failure_lookup["selected_example_grounding_issue_gate"]["level"]), "pass")
        self.assertEqual(int(failure_lookup["selected_example_grounding_issue_gate"]["value"]), 0)
        self.assertEqual(str(failure_lookup["promoted_example_coverage_gate"]["level"]), "hard_fail")


if __name__ == "__main__":
    unittest.main()
